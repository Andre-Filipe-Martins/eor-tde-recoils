#!/usr/bin/env python3
"""
igm_reionization_requirement.py
--------------------------------
Compute the total number of hydrogen-ionising photons and the corresponding
ionising energy required to reionise the diffuse IGM over the Epoch of
Reionisation (EoR), using the same cosmology and comoving volumes as the main
simulation pipeline.

The photon requirement is defined as N_gamma_req = zeta_req * N_H, where N_H
is the comoving hydrogen inventory and zeta_req = 1 + N_rec accounts for
recombinations via the case-B integral over z = 12 -> 6. The ionising-energy
requirement follows from adopting a mean photon energy <E_gamma> = 13.6 eV.

Inputs
------
This script is standalone and uses constants defined in the file.
This script computes the requirement only; it does not use the simulated TDE
catalogue.

Outputs
-------
  igm_reionization_requirement.xlsx   — formatted Excel workbook
  igm_reionization_requirement.json   — machine-readable values used by
                                        gsmf_required_for_ftde_target.py
"""

import json
import math
import os

import numpy as np
import pandas as pd
import astropy.units as u
from astropy.cosmology import FlatLambdaCDM

try:
    from openpyxl.utils import get_column_letter
    from copy import copy
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Cosmological and physical constants  (match main pipeline)
# ---------------------------------------------------------------------------
H0_KM_S_MPC = 67.74        # H0 [km/s/Mpc]
OMEGA_M      = 0.31         # matter density parameter
OMEGA_B      = 0.048        # baryon density parameter
X_HYDROGEN   = 0.75         # hydrogen mass fraction

# EoR redshift window (match main pipeline)
Z_EOR_START  = 12.0
Z_EOR_END    = 6.0
DZ_SHELL     = 0.01         # redshift integration step
FULL_SKY_SR  = 4.0 * math.pi   # full-sky solid angle [sr]

# IGM recombination model (fiducial)
IGM_CLUMPING_FACTOR  = 3.0      # C_IGM for the diffuse IGM (delta < 100)
IGM_TEMPERATURE_K    = 2.0e4    # T_IGM [K]
IGM_IONIZED_FRACTION = 1.0      # effective ionised hydrogen fraction used in the recombination integral

# Mean ionising-photon energy used to convert photons -> energy
MEAN_PHOTON_ENERGY_EV = 13.6    # eV  (Lyman-limit threshold used as the minimum photon energy)

# Physical constants
G_GRAV    = 6.67430e-11          # m^3 kg^-1 s^-2
M_PROTON  = 1.67262192369e-27    # kg
MPC_IN_M  = 3.08567758128e22     # m per Mpc
EV_TO_ERG = 1.602176634e-12      # erg per eV

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

XLSX_OUTPATH = os.path.join(BASE_DIR, "igm_reionization_requirement.xlsx")
JSON_OUTPATH = os.path.join(BASE_DIR, "igm_reionization_requirement.json")


# ---------------------------------------------------------------------------
# Cosmology
# ---------------------------------------------------------------------------
def build_cosmology():
    """Return the flat ΛCDM cosmology used by the simulation pipeline."""
    return FlatLambdaCDM(
        H0=H0_KM_S_MPC * u.km / u.s / u.Mpc,
        Om0=OMEGA_M,
        Ob0=OMEGA_B,
        Tcmb0=2.725 * u.K,
    )


# ---------------------------------------------------------------------------
# Hydrogen density
# ---------------------------------------------------------------------------
def comoving_hydrogen_density():
    """
    Compute the comoving hydrogen number density n_H [m^-3 and cm^-3].
    This is the present-day physical density expressed as a comoving density
    and is conserved in comoving coordinates at all redshifts.

    Returns
    -------
    n_H_m3  : float  — number density in m^-3
    n_H_cm3 : float  — number density in cm^-3
    """
    H0_si  = H0_KM_S_MPC * 1e3 / MPC_IN_M              # s^-1
    rho_c0 = 3.0 * H0_si**2 / (8.0 * math.pi * G_GRAV)  # kg m^-3
    rho_H0 = X_HYDROGEN * OMEGA_B * rho_c0               # kg m^-3
    n_H_m3  = rho_H0 / M_PROTON
    n_H_cm3 = n_H_m3 / 1e6                               # 1 m^3 = 1e6 cm^3
    return n_H_m3, n_H_cm3


# ---------------------------------------------------------------------------
# Case-B recombination coefficient
# ---------------------------------------------------------------------------
def alpha_B(T_K):
    """
    Case-B hydrogen recombination coefficient [cm^3 s^-1].

    Uses the standard power-law fit:
      alpha_B(T) = 2.59e-13 * (T / 10^4 K)^-0.7  cm^3 s^-1
    """
    T4 = max(T_K, 1.0) / 1.0e4
    return 2.59e-13 * T4**(-0.7)


# ---------------------------------------------------------------------------
# Midpoint redshift grid
# ---------------------------------------------------------------------------
def midpoint_z_grid(z_lo, z_hi, dz):
    """Return shell midpoints for numerical redshift integration."""
    edges = np.arange(z_lo, z_hi, dz)
    mids  = edges + 0.5 * dz
    return mids[(mids >= z_lo) & (mids <= z_hi)]


# ---------------------------------------------------------------------------
# Recombination integral -> zeta_req
# ---------------------------------------------------------------------------
def compute_zeta_req(cosmo, n_H_cm3):
    """
    Integrate the case-B recombination rate over z = Z_EOR_START -> Z_EOR_END
    to obtain the effective photon-per-hydrogen factor:
      zeta_req = 1 + N_rec
      N_rec    = integral of alpha_B * C_IGM * n_e(z)  dt

    Returns
    -------
    zeta_req : float  — photons required per hydrogen atom (>= 1)
    N_rec    : float  — mean recombinations per hydrogen atom over the EoR
    aB       : float  — case-B coefficient at IGM_TEMPERATURE_K [cm^3 s^-1]
    """
    z_grid = midpoint_z_grid(Z_EOR_END, Z_EOR_START, DZ_SHELL)

    aB = alpha_B(IGM_TEMPERATURE_K)

    # Physical hydrogen density scales as (1 + z)^3; this calculation assumes pure hydrogen for n_e.
    n_H_phys = n_H_cm3 * (1.0 + z_grid)**3      # cm^-3
    n_e_phys = IGM_IONIZED_FRACTION * n_H_phys   # cm^-3  (pure H; no HeII)

    Hz     = cosmo.H(z_grid).to(1 / u.s).value  # s^-1
    dt_dz  = 1.0 / ((1.0 + z_grid) * Hz)        # s per unit z

    rec_rate = aB * IGM_CLUMPING_FACTOR * n_e_phys  # s^-1 per H atom
    N_rec    = float(np.sum(rec_rate * dt_dz * DZ_SHELL))
    zeta_req = 1.0 + N_rec
    return zeta_req, N_rec, aB


# ---------------------------------------------------------------------------
# EoR comoving shell volume
# ---------------------------------------------------------------------------
def eor_shell_volume_Mpc3(cosmo):
    """
    Comoving volume integrated over the full-sky shell from Z_EOR_END to Z_EOR_START.
    Uses the same numerical integration as the main pipeline.

    Returns
    -------
    V_EoR_Mpc3 : float  — comoving volume [Mpc^3]
    """
    z_grid = midpoint_z_grid(Z_EOR_END, Z_EOR_START, DZ_SHELL)
    dVc_dz_dOmega = cosmo.differential_comoving_volume(z_grid).to(u.Mpc**3 / u.sr)
    V_shell = float((dVc_dz_dOmega.value * DZ_SHELL * FULL_SKY_SR).sum())
    return V_shell


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def write_results_xlsx(outpath, results_df):
    """
    Write the IGM reionization requirement results table to a formatted
    Excel workbook. openpyxl is optional; a warning is printed if unavailable.
    """
    if not _HAS_OPENPYXL:
        print(f"[warning] openpyxl not available — skipping XLSX output ({outpath})")
        print("          Install with: pip install openpyxl")
        return

    with pd.ExcelWriter(outpath, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="igm_requirement", index=False)
        wb = writer.book
        ws = wb["igm_requirement"]

        # Freeze header row and bold it
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            f = copy(cell.font)
            f.bold = True
            cell.font = f
            a = copy(cell.alignment)
            a.horizontal = "center"
            cell.alignment = a

        # Scientific notation for the Value column (column index 2)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=2).number_format = "0.000E+00"

        # Auto-fit column widths
        for j in range(1, ws.max_column + 1):
            col_letter = get_column_letter(j)
            max_len = max(
                len(str(ws.cell(row=r, column=j).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    print(f"Saved XLSX: {outpath}")




# ---------------------------------------------------------------------------
# JSON output
# ---------------------------------------------------------------------------
def write_results_json(outpath, values):
    """Write machine-readable requirement values without changing the model."""
    payload = {
        "meta": {
            "script": "igm_reionization_requirement",
            "z_eor_start": Z_EOR_START,
            "z_eor_end": Z_EOR_END,
            "dz_shell": DZ_SHELL,
            "full_sky_sr": FULL_SKY_SR,
            "H0_km_s_Mpc": H0_KM_S_MPC,
            "Omega_m": OMEGA_M,
            "Omega_b": OMEGA_B,
            "X_hydrogen": X_HYDROGEN,
            "C_IGM": IGM_CLUMPING_FACTOR,
            "T_IGM_K": IGM_TEMPERATURE_K,
            "mean_photon_energy_eV": MEAN_PHOTON_ENERGY_EV,
        },
        "results": {key: float(value) for key, value in values.items()},
    }
    with open(outpath, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
    print(f"Saved JSON: {outpath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    """Compute and export the photon-counting hydrogen-reionisation requirement."""
    cosmo = build_cosmology()

    # Hydrogen number density (comoving)
    n_H_m3, n_H_cm3 = comoving_hydrogen_density()

    # Hydrogen atoms per comoving Mpc^3
    N_H_per_Mpc3 = n_H_m3 * MPC_IN_M**3

    # Effective photon-per-hydrogen factor from the recombination integral
    zeta_req, N_rec, aB = compute_zeta_req(cosmo, n_H_cm3)

    # Photon requirement per comoving Mpc^3
    N_gamma_per_Mpc3 = zeta_req * N_H_per_Mpc3

    # Ionising-energy requirement per comoving Mpc^3
    E_photon_erg     = MEAN_PHOTON_ENERGY_EV * EV_TO_ERG
    E_ion_per_Mpc3   = N_gamma_per_Mpc3 * E_photon_erg

    # Full-sky EoR shell comoving volume
    V_EoR_Mpc3 = eor_shell_volume_Mpc3(cosmo)

    # Total photon and energy requirements over the full EoR shell
    N_gamma_EoR = N_gamma_per_Mpc3 * V_EoR_Mpc3
    E_ion_EoR   = E_ion_per_Mpc3   * V_EoR_Mpc3

    # -----------------------------------------------------------------------
    # Print the quantities used in the thesis results.
    # -----------------------------------------------------------------------
    print("\n=== IGM hydrogen reionization requirement ===\n")

    print(f"  N_rec                     = {N_rec:.3f}")
    print(f"  zeta_req (= 1 + N_rec)    = {zeta_req:.3f}  photons / H")
    print(f"  n_H (comoving)            = {N_H_per_Mpc3:.3e}  atoms Mpc^-3")
    print(f"  n_gamma_req               = {N_gamma_per_Mpc3:.3e}  photons Mpc^-3")
    print(f"  E_ion_req per Mpc^3       = {E_ion_per_Mpc3:.3e}  erg Mpc^-3")
    print(f"  V_EoR (full sky)          = {V_EoR_Mpc3:.3e}  Mpc^3")
    print(f"  N_gamma_req (EoR total)   = {N_gamma_EoR:.3e}  photons")
    print(f"  E_ion_req   (EoR total)   = {E_ion_EoR:.3e}  erg")
    print()

    # -----------------------------------------------------------------------
    # Save the requirement table to Excel.
    # -----------------------------------------------------------------------
    results_df = pd.DataFrame({
        "Quantity": [
            "N_rec",
            "zeta_req [photons/H]",
            "n_H [atoms Mpc^-3]",
            "n_gamma_req [photons Mpc^-3]",
            "E_ion_req per Mpc^3 [erg Mpc^-3]",
            "V_EoR [Mpc^3]",
            "N_gamma_req EoR total [photons]",
            "E_ion_req EoR total [erg]",
        ],
        "Value": [
            N_rec,
            zeta_req,
            N_H_per_Mpc3,
            N_gamma_per_Mpc3,
            E_ion_per_Mpc3,
            V_EoR_Mpc3,
            N_gamma_EoR,
            E_ion_EoR,
        ],
        "Notes": [
            f"Recombinations per H over z={Z_EOR_START:.0f} to {Z_EOR_END:.0f}; C_IGM={IGM_CLUMPING_FACTOR}, T={IGM_TEMPERATURE_K:.1e} K",
            "1 + N_rec; photons needed per H atom to complete reionization",
            f"Comoving; H0={H0_KM_S_MPC}, Ob={OMEGA_B}, X_H={X_HYDROGEN}",
            "zeta_req * n_H",
            f"n_gamma_req * <E_gamma>; <E_gamma> = {MEAN_PHOTON_ENERGY_EV} eV",
            f"Full-sky comoving shell z={Z_EOR_END:.0f} to {Z_EOR_START:.0f}",
            "n_gamma_req * V_EoR",
            "E_ion_req/Mpc^3 * V_EoR",
        ],
    })

    write_results_xlsx(XLSX_OUTPATH, results_df)
    write_results_json(JSON_OUTPATH, {
        "N_rec": N_rec,
        "zeta_req_photons_per_H": zeta_req,
        "n_H_atoms_per_Mpc3": N_H_per_Mpc3,
        "n_gamma_req_photons_per_Mpc3": N_gamma_per_Mpc3,
        "E_ion_req_erg_per_Mpc3": E_ion_per_Mpc3,
        "V_EoR_Mpc3": V_EoR_Mpc3,
        "N_gamma_req_EoR_total": N_gamma_EoR,
        "E_ion_req_EoR_total_erg": E_ion_EoR,
    })


if __name__ == "__main__":
    main()
