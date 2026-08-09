#!/usr/bin/env python3
"""
ratio_scan_vcent.py — Parametric scan of R = V_kick / V_cent

Scans the ratio grid R = V_kick / V_cent across all stellar-mass bins and EoR
snapshots to determine which kick velocities maximise external TDE yields.
V_cent is the minimum speed needed to reach R_cent = F_CENT * R_e:
    V_cent = sqrt( 2 * (Psi(0) - Psi(R_cent)) )
with Psi(r) = -Phi(r) the positive potential depth, normalised so Psi(0) = 0.5 * V_esc(0)^2.

The scan is run for two values of F_CENT (1.00 and 0.05) back-to-back.

Input
-----
  - ratio_scan_catalogue/runXX/data_z_*.parquet

Each catalogue row already carries its physical cap-rescaling weight.  The
scan therefore forms physical totals directly from weighted sums and never
multiplies by a separate target table.

Outputs (per F_CENT run):
  - XLSX table:  ratio_scan_vcent_tables__<fcent_tag>.xlsx
  - JSON table:  ratio_scan_vcent_rpeak__<fcent_tag>.json, including per-bin
    R_peak values used by the downstream final simulation

Combined outputs:
  - figures/external_tdes_vs_ratio_fcent_comparison.png
  - ratio_scan_vcent_figure_data.json
"""

import os

# Keep each worker single-threaded. The expensive scan is parallelised across
# snapshots below; allowing BLAS/OpenMP libraries to create extra threads in
# every worker can otherwise oversubscribe the CPU and make the run slower.
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS",      "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS",  "1")
os.environ.setdefault("OMP_NUM_THREADS",      "1")
import re
import math
import json
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from copy import copy

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import physics_relations as pr
import merger_pair_sampling as mps


# ===================== Optional openpyxl =====================
# openpyxl is required only for formatted XLSX output.
try:
    from openpyxl.utils.cell import get_column_letter
    _HAS_OPENPYXL = True
except Exception as _e:
    _HAS_OPENPYXL = False
    _OPENPYXL_ERR = f"{type(_e).__name__}: {_e}"


# ===================== PATHS =====================
try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

PARQUET_ROOT = os.path.join(BASE_DIR, "ratio_scan_catalogue")

FIG_DIR = os.path.join(BASE_DIR, "figures")
os.makedirs(FIG_DIR, exist_ok=True)

SCRIPT_STEM = "ratio_scan_vcent"
FIGURE_DATA_JSON = os.path.join(BASE_DIR, f"{SCRIPT_STEM}_figure_data.json")


# ===================== MASS BINS =====================
MASS_EDGES_ALL = np.asarray(mps.DESCENDANT_LOGM_EDGES, dtype=float)
MASS_LABELS_ALL = [mps.bin_label(MASS_EDGES_ALL[i], MASS_EDGES_ALL[i + 1])
                   for i in range(len(MASS_EDGES_ALL) - 1)]
N_MASS_ALL = len(MASS_LABELS_ALL)


# ===================== RATIO GRID =====================
# R = V_kick / V_cent
RATIO_MIN = 1.0
RATIO_MAX = 10.0
DR        = 0.10
N_R       = int(np.round((RATIO_MAX - RATIO_MIN) / DR))
RATIO_EDGES   = RATIO_MIN + DR * np.arange(N_R + 1)
RATIO_CENTERS = RATIO_EDGES[:-1] + 0.5 * DR  # ratio-bin centres used for peak selection

# Two F_CENT values scanned in sequence
F_CENT_LIST = [1.00, 0.05]

TDECAY_MODE = "vkick"   # exponential decay timescale set by V_kick

N_INT = 320             # integration resolution for representative-orbit travel times

# Physical constants / unit conversions
M_STAR_SI      = 1.3 * pr.M_sun
R_STAR_SI      = 1.3 * 6.957e8
SEC_PER_YEAR   = 365.25 * 24 * 3600
KPC_TO_KM      = 3.085677581e16   # km per kpc

F_BULGE        = 0.1548           # bulge mass fraction (Model B)

# Placeholder decay time for zero-kick events
TDECAY_NO_DECAY_YR = 1e30

# Number of independent redshift snapshots processed at the same time.
# Two workers is deliberately conservative for RAM usage. Set the environment
# variable RATIO_SCAN_WORKERS to 1 for serial execution or to a larger value
# when sufficient memory is available.
DEFAULT_WORKERS = 2
N_WORKERS = max(1, int(os.environ.get("RATIO_SCAN_WORKERS", DEFAULT_WORKERS)))


# ===================== PARQUET COLUMN DETECTION =====================
# Candidate column names for compatibility with different Parquet versions
CAND = {
    "MSTAR": ["Mstar_rem_Msun", "Mstar_Msun", "Mstar_msun", "Mstar"],
    "MBH":   ["Mrem_BH_Msun", "MBH_Msun", "Mbh_Msun", "Mbh_msun", "Mbh"],
    "RE":    ["Re_kpc", "Reff_kpc", "R_e_kpc", "R_eff_kpc", "Re"],
    "VESC":  ["Vesc0_kms", "vesc0_kms", "Vesc_kms", "vesc_kms", "Vesc0"],
    "W":     ["weight", "w", "count", "N_weight", "N", "N_galaxies", "N_targets"],
    "CLASS": ["merger_class"],
    "VERSION": ["population_model_version"],
}

try:
    import pyarrow.parquet as pq
    HAVE_PYARROW = True
except Exception:
    HAVE_PYARROW = False

NEEDED_CAND_COLS = sorted({c for lst in CAND.values() for c in lst})


def first_col(df, names):
    """Return the first name from `names` that exists as a column in `df`."""
    for n in names:
        if n in df.columns:
            return n
    return None


def parse_z_from_filename(path: str):
    """Extract the redshift value encoded in a data_z_*.parquet filename."""
    base = os.path.basename(path)
    m = re.match(r"data_z_(\d+(?:_\d+)?)\.parquet$", base)
    if not m:
        return None
    return float(m.group(1).replace("_", "."))


def discover_parquet_snapshots(parquet_root: str):
    """
    Discover snapshot Parquet files under `parquet_root`.
    Searches both root-level files and one level of subdirectories (e.g. run00/).
    Returns a sorted list of redshifts and a dict mapping z -> list of file paths.
    """
    if not os.path.isdir(parquet_root):
        raise FileNotFoundError(f"Missing parquet root dir: {parquet_root}")

    paths_by_z = {}

    for fn in os.listdir(parquet_root):
        full = os.path.join(parquet_root, fn)
        if os.path.isfile(full) and fn.startswith("data_z_") and fn.endswith(".parquet"):
            z = parse_z_from_filename(full)
            if z is not None:
                paths_by_z.setdefault(z, []).append(full)

    for entry in os.scandir(parquet_root):
        if not entry.is_dir():
            continue
        for fn in os.listdir(entry.path):
            if fn.startswith("data_z_") and fn.endswith(".parquet"):
                full = os.path.join(entry.path, fn)
                z = parse_z_from_filename(full)
                if z is not None:
                    paths_by_z.setdefault(z, []).append(full)

    if not paths_by_z:
        raise FileNotFoundError(
            f"No data_z_*.parquet found in {parquet_root} or its subdirectories."
        )

    Zs = sorted(paths_by_z.keys(), reverse=True)
    return Zs, paths_by_z


# ===================== GRAVITATIONAL POTENTIAL =====================
G_KPC = 4.30091e-6  # (km/s)^2 kpc / Msun


def psi_total_scaled_factory(rep_Mstar, rep_Re_kpc, rep_Vesc0_kms, z):
    """
    Build the positive potential depth Psi(r) = -Phi(r), normalised so that
    Psi(0) = 0.5 * V_esc(0)^2. Uses an NFW halo + Hernquist bulge model and
    rescales the model potential to match the catalogue value of Vesc0_kms.
    """
    Mh = float(pr.fire2_mh_from_mstar(rep_Mstar))
    Mb = float(F_BULGE * rep_Mstar)

    Rvir_kpc = float(pr.r_vir_mpc(Mh, z) * 1e3)
    c = float(pr.nfw_concentration(Mh, z))

    f_c = math.log(1.0 + c) - c / (1.0 + c)
    rs  = Rvir_kpc / c
    a   = rep_Re_kpc / 1.8153   # Hernquist scale length from effective radius

    def psi_nfw(r_kpc):
        r = np.asarray(r_kpc, float)
        term = np.empty_like(r)
        mask = (r > 0)
        term[mask]  = np.log1p(r[mask] / rs) / r[mask]
        term[~mask] = 1.0 / rs
        return (G_KPC * Mh / f_c) * term

    def psi_hernquist(r_kpc):
        r = np.asarray(r_kpc, float)
        return (G_KPC * Mb) / (r + a)

    psi0_model  = float(psi_nfw(0.0) + psi_hernquist(0.0))
    psi0_target = 0.5 * (rep_Vesc0_kms ** 2)
    scale = 1.0 if (not np.isfinite(psi0_model) or psi0_model <= 0) \
            else (psi0_target / psi0_model)

    def psi(r_kpc):
        return scale * (psi_nfw(r_kpc) + psi_hernquist(r_kpc))

    return psi, Rvir_kpc


def travel_time_years(psi, psi_target, r1_kpc, r2_kpc, n=N_INT):
    """
    Integrate dt = dr / sqrt(2*(Psi(r) - psi_target)) from r1 to r2 (kpc).
    Returns NaN if the integrand becomes non-physical anywhere in the interval.
    """
    if r2_kpc <= r1_kpc:
        return 0.0

    edges = np.linspace(r1_kpc, r2_kpc, n + 1)
    mids  = 0.5 * (edges[:-1] + edges[1:])
    dr    = edges[1:] - edges[:-1]

    psi_mid = psi(mids)
    diff    = psi_mid - psi_target
    if not np.all(np.isfinite(diff)) or np.any(diff <= 0):
        return np.nan

    v          = np.sqrt(2.0 * diff)         # km/s
    dt_s       = (dr / v) * KPC_TO_KM        # seconds
    dt_yr      = np.sum(dt_s) / SEC_PER_YEAR
    return float(dt_yr) if np.isfinite(dt_yr) else np.nan


def find_rmax_bisect(psi, psi_target, r_lo, r_hi):
    """
    Find the apocentric radius where Psi(r) = psi_target by bisection.
    Assumes Psi decreases monotonically with r.
    """
    flo = float(psi(r_lo) - psi_target)
    fhi = float(psi(r_hi) - psi_target)
    if not (np.isfinite(flo) and np.isfinite(fhi)):
        return None
    if flo <= 0 or fhi >= 0:
        return None

    a, b = r_lo, r_hi
    for _ in range(70):
        mid = 0.5 * (a + b)
        fm  = float(psi(mid) - psi_target)
        if not np.isfinite(fm):
            return None
        if fm > 0:
            a = mid
        else:
            b = mid
    return 0.5 * (a + b)


def build_rep_times_for_snapshot(df, z, F_CENT):
    """
    Compute representative orbit travel times per (mass bin, ratio bin) for one snapshot.

    The representative galaxy in each mass bin is taken as the median of Mstar, Re, Vesc0.
    Travel times are later scaled to individual remnants by
    (Re/Vesc0)/(Re_rep/Vesc0_rep).

    Returns
    -------
    rep_t_leave          : (N_MASS_ALL, N_R) array, years to leave central region
    rep_t_ext_max        : (N_MASS_ALL, N_R) array, full external orbit time (NaN if unbound)
    rep_scale_base       : (N_MASS_ALL,) array, Re/Vesc0 for the representative galaxy
    rep_Vcent_over_Vesc0 : (N_MASS_ALL,) array, V_cent / V_esc(0) for the representative
    """
    rep_t_leave          = np.full((N_MASS_ALL, N_R), np.nan, float)
    rep_t_ext_max        = np.full((N_MASS_ALL, N_R), np.nan, float)
    rep_scale_base       = np.full(N_MASS_ALL, np.nan, float)
    rep_Vcent_over_Vesc0 = np.full(N_MASS_ALL, np.nan, float)

    Mstar = pd.to_numeric(df["_MSTAR"], errors="coerce").to_numpy(float)
    Re    = pd.to_numeric(df["_RE"],    errors="coerce").to_numpy(float)
    Vesc0 = pd.to_numeric(df["_VESC"],  errors="coerce").to_numpy(float)

    logM   = np.log10(Mstar, where=(Mstar > 0), out=np.full_like(Mstar, np.nan))
    mass_i = np.digitize(logM, MASS_EDGES_ALL, right=False) - 1
    # np.digitize places an exact final-edge value one bin beyond the array.
    # Keep that boundary in the last descendant bin instead of dropping it.
    on_final_edge = np.isclose(logM, MASS_EDGES_ALL[-1], rtol=0.0, atol=1e-10)
    mass_i[on_final_edge] = N_MASS_ALL - 1

    for m in range(N_MASS_ALL):
        sel = (
            (mass_i == m) &
            np.isfinite(Mstar) & (Mstar > 0) &
            np.isfinite(Re)    & (Re    > 0) &
            np.isfinite(Vesc0) & (Vesc0 > 0)
        )
        if not np.any(sel):
            continue

        rep_Mstar = float(np.nanmedian(Mstar[sel]))
        rep_Re    = float(np.nanmedian(Re[sel]))
        rep_Vesc0 = float(np.nanmedian(Vesc0[sel]))
        if not (rep_Mstar > 0 and rep_Re > 0 and rep_Vesc0 > 0):
            continue

        rep_scale_base[m] = rep_Re / rep_Vesc0

        psi, Rvir_kpc = psi_total_scaled_factory(rep_Mstar, rep_Re, rep_Vesc0, z)
        R_cent = float(F_CENT) * rep_Re
        psi0   = 0.5 * rep_Vesc0 ** 2

        psi_Rc0 = float(psi(R_cent))
        if np.isfinite(psi_Rc0) and (psi0 > 0) and (psi_Rc0 < psi0):
            Vcent_over_Vesc0 = math.sqrt(max(0.0, 1.0 - psi_Rc0 / psi0))
            rep_Vcent_over_Vesc0[m] = Vcent_over_Vesc0
        else:
            rep_Vcent_over_Vesc0[m] = 1.0
            Vcent_over_Vesc0 = 1.0

        for rbin, rc in enumerate(RATIO_CENTERS):
            rc_eff     = rc * Vcent_over_Vesc0
            psi_target = psi0 * (1.0 - rc_eff ** 2)

            psi_Rc = psi_Rc0
            if not np.isfinite(psi_Rc) or psi_Rc <= psi_target:
                continue

            t_leave = travel_time_years(psi, psi_target, 0.0, R_cent)
            if not (np.isfinite(t_leave) and t_leave > 0):
                continue
            rep_t_leave[m, rbin] = t_leave

            # Orbits with psi_target <= 0 are escape-like; no turning point
            if psi_target <= 0:
                continue

            r_hi    = min(max(2.0 * R_cent, 1.2 * R_cent), Rvir_kpc)
            psi_hi  = float(psi(r_hi))
            tries   = 0
            while np.isfinite(psi_hi) and psi_hi > psi_target and (r_hi < Rvir_kpc) and tries < 25:
                r_hi   = min(r_hi * 1.6, Rvir_kpc)
                psi_hi = float(psi(r_hi))
                tries += 1

            if not (np.isfinite(psi_hi) and psi_hi < psi_target):
                continue

            rmax = find_rmax_bisect(psi, psi_target, R_cent, r_hi)
            if rmax is None or not (np.isfinite(rmax) and rmax > R_cent):
                continue

            t_out_half = travel_time_years(psi, psi_target, R_cent, rmax)
            if not (np.isfinite(t_out_half) and t_out_half > 0):
                continue

            rep_t_ext_max[m, rbin] = 2.0 * t_out_half

    return rep_t_leave, rep_t_ext_max, rep_scale_base, rep_Vcent_over_Vesc0


def integral_exp(rate_yr, tdecay_yr, t0, t1):
    """
    Integrate an exponentially decaying rate over a time interval.
    Supports NumPy broadcasting and returns zero where inputs are non-positive
    or non-finite.
    """
    r, td, t0a, t1a = np.broadcast_arrays(
        np.asarray(rate_yr,   dtype=float),
        np.asarray(tdecay_yr, dtype=float),
        np.asarray(t0,        dtype=float),
        np.asarray(t1,        dtype=float),
    )

    out = np.zeros_like(r, dtype=float)
    ok  = (
        np.isfinite(r)   & (r   > 0) &
        np.isfinite(td)  & (td  > 0) &
        np.isfinite(t0a) & np.isfinite(t1a) &
        (t1a > t0a)
    )
    if not np.any(ok):
        return out

    x0     = np.clip(t0a[ok] / td[ok], 0.0, 1e6)
    x1     = np.clip(t1a[ok] / td[ok], 0.0, 1e6)
    out[ok] = r[ok] * td[ok] * (np.exp(-x0) - np.exp(-x1))
    return np.where(np.isfinite(out) & (out > 0), out, 0.0)


# ===================== PLOT HELPERS =====================
def save_fig(fig, path_png):
    """Save a Matplotlib figure to disk."""
    fig.savefig(path_png, dpi=200, bbox_inches="tight")


def set_logy_positive(ax, y2d, pad_top=1.10, pad_bottom=1.30):
    """Set a log y-scale using only the positive finite values in y2d."""
    y   = np.asarray(y2d, float)
    pos = y[np.isfinite(y) & (y > 0)]
    if pos.size == 0:
        return False, None, None
    ymin = float(np.min(pos))
    ymax = float(np.max(pos))
    ax.set_yscale("log")
    ax.set_ylim(ymin / pad_bottom, ymax * pad_top)
    return True, ymin / pad_bottom, ymax * pad_top


def set_ratio_xticks(ax, rmin, rmax):
    """Set x-axis ticks and limits for the ratio axis."""
    ax.set_xlim(rmin, rmax)
    i0 = int(math.ceil(rmin))
    i1 = int(math.floor(rmax))
    if i1 >= i0 and (i1 - i0) >= 3:
        ax.set_xticks(np.arange(i0, i1 + 1, 1.0))
        ax.set_xticks(np.arange(i0, i1 + 0.5, 0.5), minor=True)
    else:
        ax.set_xticks(np.linspace(rmin, rmax, 6))
    ax.tick_params(axis="x", which="major", labelsize=10)


# ===================== TABLE OUTPUT HELPERS =====================
def fcent_tag(fcent: float) -> str:
    """Return a filesystem-friendly tag for F_CENT, e.g. 1.00 -> 'fcent1p00'."""
    return f"fcent{fcent:.2f}".replace(".", "p")


def write_peak_json(
    outpath,
    meta,
    peaks_rows,
    rpeak_by_bin,
    overall_best,
    rpeak_by_class,
    grids_by_class,
):
    """Save combined production peaks and class-specific diagnostics."""
    payload = {
        "meta": meta,
        # Downstream simulation.py consumes the combined population peaks.
        "R_peak_by_bin": rpeak_by_bin,
        "R_peak_by_bin_by_class": rpeak_by_class,
        "bins": peaks_rows,
        "overall_max": overall_best,
        "external_tde_grid_by_class": {
            cls: _json_safe_value(arr) for cls, arr in grids_by_class.items()
        },
    }
    with open(outpath, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"Saved peak table JSON: {outpath}")


def _json_safe_value(x):
    """
    Convert NumPy/Python values into JSON-safe values.
    Non-finite floating-point values are written as None.
    """
    if isinstance(x, np.ndarray):
        return [_json_safe_value(v) for v in x.tolist()]

    if isinstance(x, (list, tuple)):
        return [_json_safe_value(v) for v in x]

    if isinstance(x, dict):
        return {str(k): _json_safe_value(v) for k, v in x.items()}

    if isinstance(x, (np.integer,)):
        return int(x)

    if isinstance(x, (np.floating, float)):
        x = float(x)
        return x if np.isfinite(x) else None

    return x


def write_figure_data_json(outpath, results):
    """Cache combined and class-specific ratio-scan curves for later plotting."""
    payload = {
        "meta": {
            "script": SCRIPT_STEM,
            "figure": "figures/external_tdes_vs_ratio_fcent_comparison.png",
            "description": (
                "Physical external TDE totals from direct row-weighted sums. "
                "The main figure uses the combined major+minor population."
            ),
            "ratio_definition": "R = V_kick / V_cent",
            "RATIO_MIN": float(RATIO_MIN),
            "RATIO_MAX": float(RATIO_MAX),
            "DR": float(DR),
            "N_R": int(N_R),
            "TDECAY_MODE": str(TDECAY_MODE),
            "F_BULGE": float(F_BULGE),
            "N_INT": int(N_INT),
        },
        "ratio_edges": _json_safe_value(RATIO_EDGES),
        "ratio_centers": _json_safe_value(RATIO_CENTERS),
        "mass_edges_log10Mstar": _json_safe_value(MASS_EDGES_ALL),
        "mass_labels": list(MASS_LABELS_ALL),
        "fcent_order": [],
        "series": {},
    }

    for fcent, totals_by_class in results:
        key = fcent_tag(float(fcent))
        payload["fcent_order"].append(key)
        payload["series"][key] = {"F_CENT": float(fcent), "classes": {}}
        for cls, values in totals_by_class.items():
            arr = np.asarray(values, dtype=float)
            if arr.shape != (N_MASS_ALL, N_R):
                raise ValueError(
                    f"Unexpected ratio grid for {key}/{cls}: {arr.shape}; "
                    f"expected {(N_MASS_ALL, N_R)}"
                )
            payload["series"][key]["classes"][cls] = {
                "N_TDE_external_by_bin": {
                    MASS_LABELS_ALL[m]: _json_safe_value(arr[m, :])
                    for m in range(N_MASS_ALL)
                }
            }

    with open(outpath, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"Saved figure-data JSON: {outpath}")


def write_tables_xlsx(outpath, peaks_df, grid_dfs):
    """Write peak summaries and combined/major/minor ratio grids."""
    if not _HAS_OPENPYXL:
        print(f"[warning] openpyxl not available; skipping XLSX output: {outpath}")
        return

    def _format_sheet(ws, numfmt_by_colname, freeze_cell="A2"):
        ws.freeze_panes = freeze_cell
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            f = copy(cell.font); f.bold = True; cell.font = f
            a = copy(cell.alignment); a.horizontal = "center"; cell.alignment = a
        header = [c.value for c in ws[1]]
        name_to_col = {name: j + 1 for j, name in enumerate(header) if name is not None}
        for colname, fmt in numfmt_by_colname.items():
            if colname not in name_to_col:
                continue
            cidx = name_to_col[colname]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=cidx).number_format = fmt
        for j in range(1, ws.max_column + 1):
            col_letter = get_column_letter(j)
            max_len = max(
                (len(str(ws.cell(row=r, column=j).value or ""))
                 for r in range(1, ws.max_row + 1)),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    def _write(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            if peaks_df is not None and not peaks_df.empty:
                peaks_df.to_excel(writer, sheet_name="peaks", index=False)
            for cls, grid_df in grid_dfs.items():
                if grid_df is not None and not grid_df.empty:
                    grid_df.to_excel(writer, sheet_name=f"grid_{cls}"[:31], index=False)

            wb = writer.book
            if "peaks" in wb.sheetnames:
                _format_sheet(wb["peaks"], {
                    "bin_lo_log10M": "0.000",
                    "bin_hi_log10M": "0.000",
                    "R_peak_combined": "0.000",
                    "peak_total_combined": "0.00E+00",
                    "R_peak_major": "0.000",
                    "R_peak_minor": "0.000",
                    "major_at_combined_peak": "0.00E+00",
                    "minor_at_combined_peak": "0.00E+00",
                })
            for sheet in [name for name in wb.sheetnames if name.startswith("grid_")]:
                fmts = {"R_center": "0.00"}
                fmts.update({lab: "0.00E+00" for lab in MASS_LABELS_ALL})
                _format_sheet(wb[sheet], fmts)
        print(f"Saved XLSX tables: {path}")

    try:
        _write(outpath)
    except PermissionError:
        ts = time.strftime("%Y%m%d_%H%M%S")
        root, ext = os.path.splitext(outpath)
        alt = f"{root}_{ts}{ext}"
        print(f"[warning] XLSX appears open/locked. Writing to: {alt}")
        _write(alt)


# ===================== SCAN FOR ONE F_CENT VALUE =====================
# ===================== FAST SNAPSHOT SCAN =====================
def _read_and_normalise_catalogue(path):
    """Read the required catalogue columns and map them to canonical names."""
    if HAVE_PYARROW:
        schema_cols = set(pq.ParquetFile(path).schema.names)
        cols_to_read = [c for c in NEEDED_CAND_COLS if c in schema_cols]
        df = pd.read_parquet(path, columns=cols_to_read) if cols_to_read else pd.read_parquet(path)
    else:
        df = pd.read_parquet(path)

    if df is None or df.empty:
        return None

    c_mstar = first_col(df, CAND["MSTAR"])
    c_mbh = first_col(df, CAND["MBH"])
    c_re = first_col(df, CAND["RE"])
    c_vesc = first_col(df, CAND["VESC"])
    c_w = first_col(df, CAND["W"])
    c_class = first_col(df, CAND["CLASS"])
    c_version = first_col(df, CAND["VERSION"])
    required_cols = [c_mstar, c_mbh, c_re, c_vesc, c_w, c_class, c_version]
    if any(col is None for col in required_cols):
        raise KeyError(
            "The ratio-scan catalogues must contain descendant mass, remnant BH mass, "
            "Re, Vesc0, weight, merger_class and population_model_version columns. "
            "Rerun ratio_scan_sample_gen.py with the matching updated files."
        )

    versions = set(df[c_version].dropna().astype(str).unique())
    if versions != {mps.MODEL_VERSION}:
        raise RuntimeError(
            f"Incompatible catalogue model version in {os.path.basename(path)}: "
            f"found {sorted(versions)}, expected {mps.MODEL_VERSION!r}. "
            "Rerun ratio_scan_sample_gen.py before the ratio scan."
        )

    return df.rename(columns={
        c_mstar: "_MSTAR", c_mbh: "_MBH", c_re: "_RE",
        c_vesc: "_VESC", c_w: "_W", c_class: "_CLASS",
        c_version: "_MODEL_VERSION",
    })


def _representative_bins_in_catalogue(df):
    """Return descendant bins present in a catalogue for representative orbits."""
    mstar = pd.to_numeric(df["_MSTAR"], errors="coerce").to_numpy(float)
    logm = np.log10(
        mstar,
        where=(mstar > 0),
        out=np.full(len(df), np.nan, dtype=float),
    )
    bins = np.searchsorted(MASS_EDGES_ALL, logm, side="right") - 1
    bins[np.isclose(logm, MASS_EDGES_ALL[-1], rtol=0.0, atol=1e-10)] = N_MASS_ALL - 1
    present = bins[(bins >= 0) & (bins < N_MASS_ALL)]
    return np.unique(present)


def _prepare_run_arrays(df):
    """Validate rows and precompute all quantities independent of F_CENT and R."""
    Mstar = pd.to_numeric(df["_MSTAR"], errors="coerce").to_numpy(float)
    Mbh = pd.to_numeric(df["_MBH"], errors="coerce").to_numpy(float)
    Re = pd.to_numeric(df["_RE"], errors="coerce").to_numpy(float)
    Vesc0 = pd.to_numeric(df["_VESC"], errors="coerce").to_numpy(float)
    W = pd.to_numeric(df["_W"], errors="coerce").to_numpy(float)
    merger_class = df["_CLASS"].astype(str).str.lower().to_numpy()

    logM = np.log10(Mstar, where=(Mstar > 0), out=np.full_like(Mstar, np.nan))
    mass_i = np.searchsorted(MASS_EDGES_ALL, logM, side="right") - 1
    mass_i[np.isclose(logM, MASS_EDGES_ALL[-1], rtol=0.0, atol=1e-10)] = N_MASS_ALL - 1

    valid = (
        (mass_i >= 0) & (mass_i < N_MASS_ALL) &
        np.isfinite(Mbh) & (Mbh > 0) &
        np.isfinite(Re) & (Re > 0) &
        np.isfinite(Vesc0) & (Vesc0 > 0) &
        np.isfinite(W) & (W > 0) &
        np.isin(merger_class, ["major", "minor"])
    )
    if not np.any(valid):
        return None

    mass_i = mass_i[valid].astype(int)
    Mbh = Mbh[valid]
    Re = Re[valid]
    Vesc0 = Vesc0[valid]
    W = W[valid]
    merger_class = merger_class[valid]

    Mbh_SI = Mbh * pr.M_sun
    sigma_km_s, _ = pr.sigma_from_mbh(Mbh_SI)
    rt = pr.r_t(Mbh_SI, M_STAR_SI, R_STAR_SI)

    # These factors were previously recomputed for every one of the 90 ratio
    # values, although they depend only on the remnant catalogue row.
    mb_coll_prefactor_msun = (
        4.0e4
        * (Mbh / 1.0e7) ** (-0.25)
        * (sigma_km_s / 100.0) ** 2.5
    )
    ln_lambda = np.log(Mbh_SI / M_STAR_SI)
    t_decay_mass_prefactor_yr = 1.5e9 * (Mbh / 1.0e7) ** 2

    return {
        "mass_i": mass_i,
        "Mbh": Mbh,
        "Mbh_SI": Mbh_SI,
        "Re": Re,
        "Vesc0": Vesc0,
        "W": W,
        "major_mask": merger_class == "major",
        "minor_mask": merger_class == "minor",
        "sigma_km_s": sigma_km_s,
        "rt": rt,
        "mb_coll_prefactor_msun": mb_coll_prefactor_msun,
        "ln_lambda": ln_lambda,
        "t_decay_mass_prefactor_yr": t_decay_mass_prefactor_yr,
        "re_over_vesc": Re / Vesc0,
    }


def _scan_prepared_run(prep, dt_to_z6_yr, rep_data):
    """Scan one prepared run for one F_CENT value.

    The physical equations are unchanged. Work that is independent of the
    ratio is cached, rows that do not cross the boundary before z=6 are skipped,
    and the two exponential integrals reuse the same decay exponentials.
    """
    rep_t_leave, rep_t_ext_max, rep_scale_base, rep_Vcent_over_Vesc0 = rep_data

    mass_i = prep["mass_i"]
    missing_rep_bins = np.unique(mass_i[~np.isfinite(rep_scale_base[mass_i])])
    if len(missing_rep_bins):
        missing_labels = [MASS_LABELS_ALL[int(b)] for b in missing_rep_bins]
        raise RuntimeError(
            f"No valid representative descendant potential for bins {missing_labels}."
        )

    base = rep_scale_base[mass_i]
    with np.errstate(divide="ignore", invalid="ignore"):
        scale = prep["re_over_vesc"] / base
    scale = np.where(np.isfinite(scale) & (scale > 0), scale, 1.0)

    vcent_ratio = rep_Vcent_over_Vesc0[mass_i]
    vcent_ratio = np.where(
        np.isfinite(vcent_ratio) & (vcent_ratio > 0), vcent_ratio, 1.0
    )
    t_cap = np.full_like(prep["Mbh"], dt_to_z6_yr, dtype=float)

    ext_run = {
        "major": np.zeros((N_MASS_ALL, N_R), float),
        "minor": np.zeros((N_MASS_ALL, N_R), float),
        "combined": np.zeros((N_MASS_ALL, N_R), float),
    }

    for rbin, rc in enumerate(RATIO_CENTERS):
        # Boundary crossing depends only on representative orbit times and the
        # exact Re/Vesc scaling. Rows that do not leave have exactly zero
        # external TDEs, so the expensive TDE calculation is unnecessary.
        t_leave_all = rep_t_leave[mass_i, rbin] * scale
        t_leave_all = np.where(
            np.isfinite(t_leave_all) & (t_leave_all > 0), t_leave_all, np.nan
        )
        active = np.isfinite(t_leave_all) & (t_leave_all < t_cap)
        if not np.any(active):
            continue

        mi = mass_i[active]
        Mbh = prep["Mbh"][active]
        Mbh_SI = prep["Mbh_SI"][active]
        Vesc0 = prep["Vesc0"][active]
        W = prep["W"][active]
        major_mask = prep["major_mask"][active]
        minor_mask = prep["minor_mask"][active]
        rt = prep["rt"][active]

        # Preserve the original multiplication order used for V_kick.
        V_kick = rc * vcent_ratio[active] * Vesc0
        V_kick = np.where(np.isfinite(V_kick), V_kick, 0.0)
        v_k_ms = V_kick * pr.km

        rk = pr.r_k(Mbh_SI, v_k_ms)
        r_eff_pc = pr.r_eff_from_rk_gamma1(rk) / pr.pc

        Mb_coll_kg = (
            prep["mb_coll_prefactor_msun"][active]
            * (r_eff_pc / 0.1) ** 1.25
        ) * pr.M_sun
        cap = np.where(
            np.isfinite(Mb_coll_kg) & (Mb_coll_kg > 0),
            Mb_coll_kg / M_STAR_SI,
            0.0,
        )
        f_b_raw = Mb_coll_kg / Mbh_SI
        f_b = np.clip(
            np.where(np.isfinite(f_b_raw), f_b_raw, 0.0), 0.0, 1.0
        )

        rk_arr = np.asarray(rk, dtype=float)
        ln_ratio = np.where(rk_arr > rt, np.log(rk_arr / rt), np.inf)
        rate_s = (
            (prep["ln_lambda"][active] / ln_ratio)
            * (v_k_ms / rk_arr)
            * f_b
        )
        rate_yr = np.where(
            np.isfinite(rate_s) & (rate_s > 0), rate_s * SEC_PER_YEAR, 0.0
        )

        if TDECAY_MODE == "vkick":
            with np.errstate(divide="ignore", invalid="ignore", over="ignore"):
                t_decay_yr = (
                    prep["t_decay_mass_prefactor_yr"][active]
                    * (V_kick / 1.0e3) ** (-3.0)
                )
            t_decay_yr = np.where(
                (V_kick <= 0) & np.isfinite(Mbh),
                TDECAY_NO_DECAY_YR,
                t_decay_yr,
            )
        elif TDECAY_MODE == "mass_only":
            t_decay_yr = prep["t_decay_mass_prefactor_yr"][active]
        else:
            raise ValueError(f"Unknown TDECAY_MODE: {TDECAY_MODE}")
        t_decay_yr = np.where(
            np.isfinite(t_decay_yr) & (t_decay_yr > 0), t_decay_yr, 0.0
        )

        t_cent_pre = t_leave_all[active]
        rem_time = np.maximum(0.0, t_cap[active] - t_cent_pre)
        t_ext_max = rep_t_ext_max[mi, rbin] * scale[active]
        t_ext = np.where(
            np.isfinite(t_ext_max) & (t_ext_max > 0),
            np.minimum(t_ext_max, rem_time),
            rem_time,
        )

        # This is the same pair of exponential integrals as integral_exp(),
        # but exp(-t_cent/t_decay) is evaluated once and reused.
        ok = (
            np.isfinite(rate_yr) & (rate_yr > 0) &
            np.isfinite(t_decay_yr) & (t_decay_yr > 0) &
            np.isfinite(t_cent_pre) & np.isfinite(t_ext)
        )
        N_cent_unc = np.zeros_like(rate_yr)
        N_ext_unc = np.zeros_like(rate_yr)
        if np.any(ok):
            amp = rate_yr[ok] * t_decay_yr[ok]
            x_pre = np.clip(t_cent_pre[ok] / t_decay_yr[ok], 0.0, 1e6)
            x_end = np.clip(
                (t_cent_pre[ok] + t_ext[ok]) / t_decay_yr[ok], 0.0, 1e6
            )
            e_pre = np.exp(-x_pre)
            e_end = np.exp(-x_end)
            cent_values = amp * (1.0 - e_pre)
            ext_values = amp * (e_pre - e_end)
            N_cent_unc[ok] = np.where(
                np.isfinite(cent_values) & (cent_values > 0), cent_values, 0.0
            )
            N_ext_unc[ok] = np.where(
                (t_ext[ok] > 0) & np.isfinite(ext_values) & (ext_values > 0),
                ext_values,
                0.0,
            )

        N_cent = np.minimum(N_cent_unc, cap)
        remaining = np.maximum(0.0, cap - N_cent)
        clipped_ext = np.minimum(N_ext_unc, remaining)
        N_ext = np.where(
            np.isfinite(clipped_ext), np.maximum(0.0, clipped_ext), 0.0
        )

        weighted = W * N_ext
        ext_run["combined"][:, rbin] = np.bincount(
            mi, weights=weighted, minlength=N_MASS_ALL
        )
        if np.any(major_mask):
            ext_run["major"][:, rbin] = np.bincount(
                mi[major_mask], weights=weighted[major_mask], minlength=N_MASS_ALL
            )
        if np.any(minor_mask):
            ext_run["minor"][:, rbin] = np.bincount(
                mi[minor_mask], weights=weighted[minor_mask], minlength=N_MASS_ALL
            )

    return ext_run


def process_snapshot_all_fcent(z, paths):
    """Process one redshift snapshot for all F_CENT values in one file pass."""
    dt_to_z6_yr = float(pr.time_until_z6(z))
    empty = {
        float(f): {
            cls: np.zeros((N_MASS_ALL, N_R), float)
            for cls in ("major", "minor", "combined")
        }
        for f in F_CENT_LIST
    }
    if not np.isfinite(dt_to_z6_yr) or dt_to_z6_yr <= 0.0:
        return {"z": float(z), "used_runs": 0, "means": empty}

    rep_by_fcent = {
        float(f): (
            np.full((N_MASS_ALL, N_R), np.nan, float),
            np.full((N_MASS_ALL, N_R), np.nan, float),
            np.full(N_MASS_ALL, np.nan, float),
            np.full(N_MASS_ALL, np.nan, float),
        )
        for f in F_CENT_LIST
    }
    run_totals = {
        float(f): {cls: [] for cls in ("major", "minor", "combined")}
        for f in F_CENT_LIST
    }

    for path in paths:
        try:
            df = _read_and_normalise_catalogue(path)
        except (KeyError, RuntimeError):
            # Preserve the original fail-fast behaviour for incompatible or
            # incomplete catalogues; silently skipping these could bias totals.
            raise
        except Exception as exc:
            print(f"  [warn] could not read {os.path.basename(path)}; skipping ({exc})")
            continue
        if df is None or df.empty:
            continue

        present_bins = _representative_bins_in_catalogue(df)
        for fcent in F_CENT_LIST:
            fkey = float(fcent)
            rep_t_leave, rep_t_ext_max, rep_scale_base, rep_vcent = rep_by_fcent[fkey]
            missing_present = [b for b in present_bins if not np.isfinite(rep_scale_base[b])]
            if missing_present:
                cand_leave, cand_ext, cand_scale, cand_vcent = build_rep_times_for_snapshot(
                    df, z, fcent
                )
                for b in missing_present:
                    if np.isfinite(cand_scale[b]):
                        rep_t_leave[b, :] = cand_leave[b, :]
                        rep_t_ext_max[b, :] = cand_ext[b, :]
                        rep_scale_base[b] = cand_scale[b]
                        rep_vcent[b] = cand_vcent[b]

        prep = _prepare_run_arrays(df)
        if prep is None:
            continue

        for fcent in F_CENT_LIST:
            fkey = float(fcent)
            ext_run = _scan_prepared_run(
                prep, dt_to_z6_yr, rep_by_fcent[fkey]
            )
            for cls in run_totals[fkey]:
                run_totals[fkey][cls].append(ext_run[cls])

    used_runs = len(run_totals[float(F_CENT_LIST[0])]["combined"])
    means = empty
    if used_runs:
        for fcent in F_CENT_LIST:
            fkey = float(fcent)
            for cls in means[fkey]:
                stack = np.stack(run_totals[fkey][cls], axis=0)
                mean_z = np.mean(stack, axis=0)
                means[fkey][cls] = np.where(
                    np.isfinite(mean_z) & (mean_z >= 0), mean_z, 0.0
                )

    return {"z": float(z), "used_runs": used_runs, "means": means}


# ===================== OUTPUT FINALISATION =====================
def finalize_one_fcent(F_CENT, totals, n_runs_nominal):
    """Extract peaks and write the same JSON/XLSX outputs as the original scan."""
    tag_fcent = fcent_tag(F_CENT)
    xlsx_out = os.path.join(BASE_DIR, f"{SCRIPT_STEM}_tables__{tag_fcent}.xlsx")
    json_out = os.path.join(BASE_DIR, f"{SCRIPT_STEM}_rpeak__{tag_fcent}.json")

    if not np.allclose(
        totals["combined"], totals["major"] + totals["minor"],
        rtol=1e-10, atol=1e-6,
    ):
        raise RuntimeError("Combined ratio-scan grid does not equal major + minor totals.")

    rpeak_by_class = {cls: {} for cls in totals}
    peaks_rows = []
    overall_best = {"val": -np.inf, "R": np.nan, "bin": None}

    print("\n" + "=" * 86)
    print(f"[ratio_scan_vcent] RESULTS: F_CENT={F_CENT:.2f} ({tag_fcent})")
    print("=" * 86)

    for m, lab in enumerate(MASS_LABELS_ALL):
        peaks = {}
        indices = {}
        for cls in totals:
            row = totals[cls][m]
            if np.any(np.isfinite(row)) and np.nanmax(row) > 0:
                idx = int(np.nanargmax(row))
                peaks[cls] = float(RATIO_CENTERS[idx])
                indices[cls] = idx
            else:
                peaks[cls] = float(RATIO_CENTERS[0])
                indices[cls] = 0
            rpeak_by_class[cls][lab] = peaks[cls]

        j_comb = indices["combined"]
        combined_peak = float(totals["combined"][m, j_comb])
        major_at_combined = float(totals["major"][m, j_comb])
        minor_at_combined = float(totals["minor"][m, j_comb])
        peaks_rows.append({
            "bin_label": lab,
            "bin_lo_log10M": float(MASS_EDGES_ALL[m]),
            "bin_hi_log10M": float(MASS_EDGES_ALL[m + 1]),
            "R_peak_combined": peaks["combined"],
            "peak_total_combined": combined_peak,
            "R_peak_major": peaks["major"],
            "R_peak_minor": peaks["minor"],
            "major_at_combined_peak": major_at_combined,
            "minor_at_combined_peak": minor_at_combined,
        })
        print(
            f"  {lab}: combined R_peak={peaks['combined']:.3f}, "
            f"N_ext={combined_peak:.6g} "
            f"(major={major_at_combined:.6g}, minor={minor_at_combined:.6g})"
        )
        if combined_peak > overall_best["val"]:
            overall_best = {"val": combined_peak, "R": peaks["combined"], "bin": lab}

    meta = {
        "script": SCRIPT_STEM,
        "ratio_definition": "R = V_kick / V_cent",
        "F_CENT": float(F_CENT),
        "TDECAY_MODE": str(TDECAY_MODE),
        "RATIO_MIN": float(RATIO_MIN),
        "RATIO_MAX": float(RATIO_MAX),
        "DR": float(DR),
        "R_centers": [float(x) for x in RATIO_CENTERS],
        "mass_bin_meaning": "descendant galaxy stellar mass",
        "production_peak_population": "combined major+minor",
        "weighting": "direct sum(weight * N_TDE); no second target multiplication",
        "parquet_root": PARQUET_ROOT,
        "nominal_runs": int(n_runs_nominal),
        "optimisation": (
            "shared F_CENT file pass; cached ratio-invariant row quantities; "
            "early zero-external-row skip; deterministic snapshot parallelism"
        ),
        "workers": int(N_WORKERS),
    }

    peaks_df = pd.DataFrame(peaks_rows)
    grid_dfs = {}
    for cls, arr in totals.items():
        grid_df = pd.DataFrame({"R_center": RATIO_CENTERS})
        for m, lab in enumerate(MASS_LABELS_ALL):
            grid_df[lab] = arr[m]
        grid_dfs[cls] = grid_df

    write_peak_json(
        json_out, meta, peaks_rows, rpeak_by_class["combined"], overall_best,
        rpeak_by_class, totals,
    )
    write_tables_xlsx(xlsx_out, peaks_df, grid_dfs)
    print(f"[ratio_scan_vcent] DONE: F_CENT={F_CENT:.2f}")
    return totals


# ===================== MAIN =====================
def main():
    """Run both boundary scans with shared I/O and optional snapshot parallelism."""
    plt.rcParams.update({
        "font.size":      12,
        "axes.titlesize": 14,
        "axes.labelsize": 12,
        "legend.fontsize": 10,
    })

    Zs, paths_by_z = discover_parquet_snapshots(PARQUET_ROOT)
    n_runs_nominal = max(len(v) for v in paths_by_z.values())
    workers = min(N_WORKERS, len(Zs))

    print("\n" + "=" * 86)
    print("[ratio_scan_vcent] FAST SCAN")
    print(f"  snapshots       : {len(Zs)}")
    print(f"  nominal runs    : {n_runs_nominal}")
    print(f"  F_CENT values   : {F_CENT_LIST}")
    print(f"  worker processes: {workers}")
    print("=" * 86)

    snapshot_results = {}
    if workers == 1:
        for z in Zs:
            print(f"\n=== [z={z:.2f}] {len(paths_by_z[z])} run file(s) ===")
            result = process_snapshot_all_fcent(z, paths_by_z[z])
            snapshot_results[float(z)] = result
            print(f"  -> used {result['used_runs']} run(s); completed both F_CENT scans")
    else:
        with ProcessPoolExecutor(max_workers=workers) as executor:
            future_to_z = {
                executor.submit(process_snapshot_all_fcent, z, paths_by_z[z]): float(z)
                for z in Zs
            }
            for future in as_completed(future_to_z):
                z = future_to_z[future]
                result = future.result()
                snapshot_results[z] = result
                print(
                    f"  [z={z:.2f}] used {result['used_runs']} run(s); "
                    "completed both F_CENT scans"
                )

    totals_all = {
        float(f): {
            cls: np.zeros((N_MASS_ALL, N_R), float)
            for cls in ("major", "minor", "combined")
        }
        for f in F_CENT_LIST
    }

    # Add snapshots in the same descending-redshift order as the original
    # serial script, independent of the order in which workers finished.
    for z in Zs:
        result = snapshot_results.get(float(z))
        if result is None:
            continue
        for fcent in F_CENT_LIST:
            fkey = float(fcent)
            for cls in totals_all[fkey]:
                totals_all[fkey][cls] += result["means"][fkey][cls]

    results = []
    for fcent in F_CENT_LIST:
        fkey = float(fcent)
        finalized = finalize_one_fcent(fcent, totals_all[fkey], n_runs_nominal)
        results.append((fcent, finalized))

    write_figure_data_json(FIGURE_DATA_JSON, results)

    # ===================== COMBINED FIGURE =====================
    out_png = os.path.join(FIG_DIR, "external_tdes_vs_ratio_fcent_comparison.png")

    fig, axes = plt.subplots(1, 2, figsize=(15.8, 5.3), sharex=True)

    ext_tde_plots = [
        np.where(
            np.isfinite(totals_by_class["combined"]) & (totals_by_class["combined"] > 0),
            totals_by_class["combined"],
            np.nan,
        )
        for _, totals_by_class in results
    ]

    all_pos = np.concatenate([p[np.isfinite(p)] for p in ext_tde_plots]) if ext_tde_plots else np.array([])
    all_pos = all_pos[all_pos > 0]
    if all_pos.size > 0:
        ymin = float(np.min(all_pos)) / 1.30
        ymax = float(np.max(all_pos)) * 1.10
        use_common_lim = True
    else:
        use_common_lim = False

    for ax, (fcent, _), ext_plot in zip(axes, results, ext_tde_plots):
        for m in range(N_MASS_ALL):
            y = ext_plot[m, :]
            line, = ax.plot(RATIO_CENTERS, y, lw=2.2, label=MASS_LABELS_ALL[m])

            finite = np.isfinite(y)
            if np.count_nonzero(finite) == 1:
                ax.scatter(
                    RATIO_CENTERS[finite],
                    y[finite],
                    s=35,
                    color=line.get_color(),
                    marker="o",
                    zorder=5,
                )
        ax.axvline(1.0, ls="--", lw=1.5, alpha=0.8)
        ax.set_title(
            f"Total expected external TDEs\n"
            r"$f_{\rm cent}=" + f"{fcent:.2f}$"
        )
        ax.set_xlabel(r"$\mathcal{R} = V_{\rm kick} / V_{\rm cent}$")
        set_ratio_xticks(ax, RATIO_MIN, RATIO_MAX)
        ax.grid(True, which="both", ls=":", alpha=0.7)

        if use_common_lim:
            ax.set_yscale("log")
            ax.set_ylim(ymin, ymax)
        else:
            ok, _, _ = set_logy_positive(ax, ext_plot)
            if not ok:
                print(
                    f"[plot] Warning: no positive external TDE totals for "
                    f"fcent={fcent:.2f}; leaving y-axis linear."
                )

    axes[0].set_ylabel(r"$N_{\rm TDE}$")

    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(
        handles, labels,
        title="Descendant mass [log10 M*]",
        bbox_to_anchor=(1.01, 0.98),
        loc="upper left",
    )
    fig.tight_layout()
    save_fig(fig, out_png)
    plt.close(fig)

    print(f"\nSaved figure: {out_png}")


if __name__ == "__main__":
    main()
