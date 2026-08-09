#!/usr/bin/env python3
"""
external_tde_ionizing_energy.py
--------------------------------
Compute the hydrogen-ionising energy budget from external TDEs in the revised
major+minor galaxy-pair simulation.

Every simulation row already carries its single physical cap-rescaling weight,
weight = n_phys / n_samp.  This script therefore forms physical totals directly
from sum(weight * quantity) in each run and then averages those physical totals
over the independent Monte Carlo runs.  It never multiplies by the old target
table a second time.

Inputs
------
  simulation_results/runXX/data_z_*.parquet

Outputs
-------
  external_tde_ionizing_energy.json
  external_tde_ionizing_energy_series.json
  external_tde_ionizing_energy.xlsx
  figures/external_tde_ionizing_energy_per_bin.png
  figures/external_tde_ionizing_energy_total.png
"""

import json
import os
import re
from copy import copy
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

import merger_pair_sampling as mps

# Use a non-interactive backend so figures can be written on headless systems.
os.environ.setdefault("MPLBACKEND", "Agg")
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

PARQUET_DIR = os.path.join(BASE_DIR, "simulation_results")
FIG_DIR     = os.path.join(BASE_DIR, "figures")
os.makedirs(FIG_DIR, exist_ok=True)

SCRIPT_STEM  = os.path.splitext(os.path.basename(__file__))[0] if "__file__" in globals() \
               else "external_tde_ionizing_energy"
JSON_OUTPATH        = os.path.join(BASE_DIR, f"{SCRIPT_STEM}.json")
SERIES_JSON_OUTPATH = os.path.join(BASE_DIR, f"{SCRIPT_STEM}_series.json")
XLSX_OUTPATH        = os.path.join(BASE_DIR, f"{SCRIPT_STEM}.xlsx")


# ---------------------------------------------------------------------------
# Shared descendant stellar-mass bins
# ---------------------------------------------------------------------------
EDGES = np.asarray(mps.DESCENDANT_LOGM_EDGES, dtype=float)
ALL_LABELS = [mps.bin_label(EDGES[i], EDGES[i + 1])
              for i in range(len(EDGES) - 1)]
NBINS = len(ALL_LABELS)
LOGM_MIN = float(EDGES[0])
LOGM_MAX = float(EDGES[-1])
BIN_WIDTH = float(EDGES[1] - EDGES[0])

# If True, keep only rows with t_external_yr > 0.  This does not normally
# change the result because rows with no external interval already have zero
# tde_external_post.
REQUIRE_EXTERNAL_WINDOW = False

PRINT_FILE_PATHS = False
PRINT_PER_Z_MEANS = False

# ---------------------------------------------------------------------------
# Physics knobs for ionising energy per TDE
# ---------------------------------------------------------------------------
M_STAR_TDE_MSUN = 1.3   # disrupted stellar mass [M_sun]
R_STAR_TDE_RSUN = 1.3   # disrupted stellar radius [R_sun]
F_DISK          = 0.5   # fraction of stellar debris that forms the accretion disc
F_ION_PHASE     = 1.0   # fixed-energy path fraction going into ionising radiation
ETA_FIXED       = 0.10  # fixed radiative efficiency (used when spin is unavailable)

USE_SPIN_EFFICIENCY = False  # set True only if a_spin column is present in Parquets

# BH-mass-dependent accreted mass using the Mummery analytic viscous disc model
USE_MUMMERY_MBH_DEPENDENCE = True

# Mummery model parameters
MUM_N_DECAY      = 1.2
MUM_XP           = 10.0
MUM_V_NUIS       = 1000.0
MUM_TION_NORM_YR = 200.0
MUM_MBH_REF_MSUN = 1.0e6

# Diagnostic Rydberg-equivalent photon count: N_equiv = E_ion / (13.6 eV).
PRINT_EQUIV_PHOTONS = True
E_PHOTON_EFF_EV     = 13.6


# ---------------------------------------------------------------------------
# Physical constants
# ---------------------------------------------------------------------------
C_LIGHT   = 299_792_458.0   # speed of light [m/s]
M_SUN     = 1.98847e30      # solar mass [kg]
EV_TO_ERG = 1.602176634e-12 # 1 eV in erg
G_GRAV    = 6.67430e-11     # gravitational constant [SI]
R_SUN     = 6.957e8         # solar radius [m]
YR_TO_S   = 365.25 * 24 * 3600.0


# ---------------------------------------------------------------------------
# Plot colours
# ---------------------------------------------------------------------------
# Consistent plotting colour assigned to each host-mass bin.
_default_colors = plt.rcParams["axes.prop_cycle"].by_key().get("color", [])
COLOR_BY_LABEL  = {lbl: _default_colors[j % len(_default_colors)]
                   for j, lbl in enumerate(ALL_LABELS)}


# ---------------------------------------------------------------------------
# Snapshot discovery
# ---------------------------------------------------------------------------
def _parse_redshift_from_filename(fname: str) -> Optional[float]:
    """Extract the redshift encoded in a data_z_*.parquet filename."""
    match = re.match(r"data_z_(\d+(?:_\d+)?)\.parquet$", fname)
    if not match:
        return None
    z_str = match.group(1).replace("_", ".")
    try:
        return round(float(z_str), 2)
    except ValueError:
        return None


def discover_parquet_snapshots(
    parquet_dir: str,
) -> Tuple[np.ndarray, Dict[float, List[str]]]:
    """
    Walk parquet_dir for data_z_*.parquet files at root level and in run
    subdirectories. Both legacy root-level files and current runXX/
    subdirectories are supported. Returns a sorted array of redshifts and a dict mapping
    each redshift to the list of file paths for that snapshot.
    """
    if not os.path.isdir(parquet_dir):
        raise SystemExit(f"[ERROR] Missing simulation results directory: {parquet_dir}")

    paths_by_z: Dict[float, List[str]] = {}

    # Root-level files (legacy layout)
    for fname in os.listdir(parquet_dir):
        full = os.path.join(parquet_dir, fname)
        if os.path.isfile(full) and fname.startswith("data_z_") and fname.endswith(".parquet"):
            z = _parse_redshift_from_filename(fname)
            if z is not None:
                paths_by_z.setdefault(z, []).append(full)

    # Per-run subdirectories (current layout: simulation_results/<run_tag>/)
    for entry in os.scandir(parquet_dir):
        if not entry.is_dir():
            continue
        for fname in os.listdir(entry.path):
            if not (fname.startswith("data_z_") and fname.endswith(".parquet")):
                continue
            z = _parse_redshift_from_filename(fname)
            if z is not None:
                paths_by_z.setdefault(z, []).append(os.path.join(entry.path, fname))

    if not paths_by_z:
        raise SystemExit("[ERROR] No Parquet snapshot files found in simulation_results/.")

    for paths in paths_by_z.values():
        paths.sort()
    z_array = np.array(sorted(paths_by_z.keys(), reverse=True), dtype=float)
    print("Parquet directory:", parquet_dir)
    print("Redshift snapshots found:", ", ".join(f"{z:.2f}" for z in z_array))
    return z_array, paths_by_z


# ---------------------------------------------------------------------------
# Physics: ionising energy per TDE
# ---------------------------------------------------------------------------
def ionizing_energy_per_tde_erg(eta: float = ETA_FIXED) -> float:
    """
    Return the ionising energy per TDE for the fixed-accreted-mass path.
    """
    delta_m_acc_kg = M_STAR_TDE_MSUN * M_SUN * F_DISK * F_ION_PHASE
    return float(eta * delta_m_acc_kg * C_LIGHT**2 * 1e7)  # J -> erg


def _mummery_viscous_timescale_sec(v_nuis: float = MUM_V_NUIS) -> float:
    """Viscous timescale t_visc for the Mummery analytic disc model [seconds]."""
    m_star = M_STAR_TDE_MSUN * M_SUN
    r_star = R_STAR_TDE_RSUN * R_SUN
    t_star = np.sqrt(8.0 * r_star**3 / (G_GRAV * m_star))
    return float(v_nuis * t_star)


def _mummery_ionizing_timescale_sec(
    mbh_msun: np.ndarray,
    v_nuis: float = MUM_V_NUIS,
    n_decay: float = MUM_N_DECAY,
    x_peak: float = MUM_XP,
) -> np.ndarray:
    """
    BH-mass-dependent ionising timescale t_ion [seconds] from Mummery et al.,
    parameterised by the disc viscosity nuisance parameter v_nuis.
    """
    mbh = np.asarray(mbh_msun, dtype=float)
    mbh = np.where(np.isfinite(mbh) & (mbh > 0.0), mbh, np.nan)
    t_yr = (
        MUM_TION_NORM_YR
        * (v_nuis / 1000.0)
        * (mbh / MUM_MBH_REF_MSUN) ** (-1.0 / n_decay)
        * (x_peak / 10.0) ** (-2.0 / n_decay)
    )
    return t_yr * YR_TO_S


def mummery_accreted_mass_msun(
    mbh_msun: np.ndarray,
    m_disk_msun: float,
    n_decay: float = MUM_N_DECAY,
    x_peak: float = MUM_XP,
    v_nuis: float = MUM_V_NUIS,
) -> np.ndarray:
    """
    Accreted disc mass [M_sun] as a function of BH mass, using the Mummery
    analytic viscous-disc solution.  The returned fraction is clipped to [0, 1].
    """
    t_visc = _mummery_viscous_timescale_sec(v_nuis=v_nuis)
    t_ion  = _mummery_ionizing_timescale_sec(mbh_msun, v_nuis=v_nuis, n_decay=n_decay, x_peak=x_peak)

    ratio = t_ion / t_visc
    frac  = np.where(ratio > 1.0, 1.0 - ratio ** (1.0 - n_decay), 0.0)
    frac  = np.clip(frac, 0.0, 1.0)
    frac  = np.where(np.isfinite(frac), frac, 0.0)
    return float(m_disk_msun) * frac


# ---------------------------------------------------------------------------
# Main aggregation: direct physical row-weighted sums
# ---------------------------------------------------------------------------
CLASSES = (mps.MERGER_MAJOR, mps.MERGER_MINOR, "combined")


def _descendant_bin_indices(mstar_msun: np.ndarray) -> np.ndarray:
    """Assign rows to the shared descendant bins, including the final edge."""
    mstar = np.asarray(mstar_msun, dtype=float)
    logm = np.log10(mstar, where=(mstar > 0.0), out=np.full_like(mstar, np.nan))
    idx = np.searchsorted(EDGES, logm, side="right") - 1
    idx[np.isclose(logm, EDGES[-1], rtol=0.0, atol=1.0e-10)] = NBINS - 1
    return idx.astype(np.int16, copy=False)


def _read_weighted_columns(path: str) -> pd.DataFrame:
    """Read only the columns needed for the weighted energy aggregation."""
    columns = [
        "population_model_version", "merger_class", "weight",
        "Mstar_rem_Msun", "Mrem_BH_Msun", "tde_external_post",
        "t_external_yr",
    ]
    if USE_SPIN_EFFICIENCY:
        columns.append("a_spin")
    try:
        df = pd.read_parquet(path, columns=columns)
    except Exception:
        df = pd.read_parquet(path)

    required = {
        "population_model_version", "merger_class", "weight",
        "Mstar_rem_Msun", "Mrem_BH_Msun", "tde_external_post",
    }
    missing = sorted(required.difference(df.columns))
    if missing:
        raise KeyError(
            f"Incomplete revised simulation catalogue {path}: missing {missing}. "
            "Rerun simulation.py with the matching updated files."
        )

    versions = set(df["population_model_version"].dropna().astype(str).unique())
    if versions != {mps.MODEL_VERSION}:
        raise RuntimeError(
            f"Incompatible population model in {path}: found {sorted(versions)}, "
            f"expected {mps.MODEL_VERSION!r}."
        )
    return df


def accumulate_external_tdes_and_energy(
    z_array: np.ndarray,
    paths_by_z: Dict[float, List[str]],
):
    """Average direct physical weighted totals over runs at every redshift.

    Returns
    -------
    z_plot, combined N/E series used by the figures, and a class_series mapping
    containing combined, major, and minor physical totals.
    """
    n_z = len(z_array)
    n_by_class = {cls: np.zeros((n_z, NBINS), dtype=float) for cls in CLASSES}
    e_by_class = {cls: np.zeros((n_z, NBINS), dtype=float) for cls in CLASSES}
    eion_fixed = ionizing_energy_per_tde_erg(ETA_FIXED)

    for iz, z in enumerate(z_array):
        paths = paths_by_z[z]
        print(f"\n[z = {z:.2f}] Direct weighted aggregation over {len(paths)} file(s)")
        if PRINT_FILE_PATHS:
            for path in paths:
                print("  ", path)

        run_n = {cls: [] for cls in CLASSES}
        run_e = {cls: [] for cls in CLASSES}

        for path in paths:
            df = _read_weighted_columns(path)
            if df is None or df.empty:
                continue

            if REQUIRE_EXTERNAL_WINDOW and "t_external_yr" in df.columns:
                t_ext = pd.to_numeric(df["t_external_yr"], errors="coerce").fillna(0.0).to_numpy(float)
                df = df.loc[t_ext > 0.0].copy()
                if df.empty:
                    continue

            mstar = pd.to_numeric(df["Mstar_rem_Msun"], errors="coerce").to_numpy(float)
            mbh = pd.to_numeric(df["Mrem_BH_Msun"], errors="coerce").to_numpy(float)
            n_ext = pd.to_numeric(df["tde_external_post"], errors="coerce").fillna(0.0).to_numpy(float)
            weight = pd.to_numeric(df["weight"], errors="coerce").fillna(0.0).to_numpy(float)
            merger_class = df["merger_class"].astype(str).str.lower().to_numpy()
            bin_idx = _descendant_bin_indices(mstar)

            valid = (
                (bin_idx >= 0) & (bin_idx < NBINS)
                & np.isfinite(mbh) & (mbh > 0.0)
                & np.isfinite(weight) & (weight > 0.0)
                & np.isin(merger_class, [mps.MERGER_MAJOR, mps.MERGER_MINOR])
            )
            if not np.any(valid):
                continue

            bin_idx = bin_idx[valid].astype(int)
            mbh = mbh[valid]
            weight = weight[valid]
            merger_class = merger_class[valid]
            n_ext = np.where(np.isfinite(n_ext[valid]) & (n_ext[valid] > 0.0), n_ext[valid], 0.0)

            if USE_SPIN_EFFICIENCY and "a_spin" in df.columns:
                spin = pd.to_numeric(df["a_spin"], errors="coerce").to_numpy(float)[valid]
                spin = np.clip(spin, 0.0, 0.998)
                eta = np.where(np.isfinite(spin), 0.057 + 0.3 * spin, ETA_FIXED)
            else:
                eta = np.full(len(mbh), ETA_FIXED, dtype=float)

            if USE_MUMMERY_MBH_DEPENDENCE:
                m_disk = M_STAR_TDE_MSUN * F_DISK
                delta_m_acc = mummery_accreted_mass_msun(mbh, m_disk)
                eion_per_tde = eta * (delta_m_acc * M_SUN) * C_LIGHT**2 * 1.0e7
                eion_per_tde = np.where(
                    np.isfinite(eion_per_tde) & (eion_per_tde > 0.0),
                    eion_per_tde,
                    0.0,
                )
            else:
                eion_per_tde = np.full(len(mbh), eion_fixed, dtype=float)

            weighted_n = weight * n_ext
            weighted_e = weighted_n * eion_per_tde

            per_run_n = {}
            per_run_e = {}
            for cls in (mps.MERGER_MAJOR, mps.MERGER_MINOR):
                mask = merger_class == cls
                per_run_n[cls] = np.bincount(
                    bin_idx[mask], weights=weighted_n[mask], minlength=NBINS
                ).astype(float)
                per_run_e[cls] = np.bincount(
                    bin_idx[mask], weights=weighted_e[mask], minlength=NBINS
                ).astype(float)
            per_run_n["combined"] = per_run_n[mps.MERGER_MAJOR] + per_run_n[mps.MERGER_MINOR]
            per_run_e["combined"] = per_run_e[mps.MERGER_MAJOR] + per_run_e[mps.MERGER_MINOR]

            for cls in CLASSES:
                run_n[cls].append(per_run_n[cls])
                run_e[cls].append(per_run_e[cls])

        used_runs = len(run_n["combined"])
        if used_runs == 0:
            print("  [warning] no valid revised catalogues at this redshift")
            continue

        for cls in CLASSES:
            n_by_class[cls][iz] = np.mean(np.stack(run_n[cls], axis=0), axis=0)
            e_by_class[cls][iz] = np.mean(np.stack(run_e[cls], axis=0), axis=0)

        if not np.allclose(
            n_by_class["combined"][iz],
            n_by_class[mps.MERGER_MAJOR][iz] + n_by_class[mps.MERGER_MINOR][iz],
            rtol=1.0e-10, atol=1.0e-6,
        ):
            raise RuntimeError(f"Major+minor TDE audit failed at z={z:.2f}")
        if not np.allclose(
            e_by_class["combined"][iz],
            e_by_class[mps.MERGER_MAJOR][iz] + e_by_class[mps.MERGER_MINOR][iz],
            rtol=1.0e-10, atol=1.0e40,
        ):
            raise RuntimeError(f"Major+minor energy audit failed at z={z:.2f}")

        if PRINT_PER_Z_MEANS:
            print(
                f"  mean physical totals: N_ext={n_by_class['combined'][iz].sum():.6e}, "
                f"E_ion={e_by_class['combined'][iz].sum():.6e} erg"
            )

    order = np.argsort(z_array)
    z_plot = np.asarray(z_array, dtype=float)[order]
    class_series = {}
    for cls in CLASSES:
        n_ordered = n_by_class[cls][order]
        e_ordered = e_by_class[cls][order]
        class_series[cls] = {
            "N_ext_by_bin": {
                label: n_ordered[:, ibin].copy() for ibin, label in enumerate(ALL_LABELS)
            },
            "E_ion_by_bin": {
                label: e_ordered[:, ibin].copy() for ibin, label in enumerate(ALL_LABELS)
            },
            "overall_N_ext": np.sum(n_ordered, axis=1),
            "overall_E_ion": np.sum(e_ordered, axis=1),
        }

    combined = class_series["combined"]
    return (
        z_plot,
        combined["N_ext_by_bin"],
        combined["E_ion_by_bin"],
        combined["overall_N_ext"],
        combined["overall_E_ion"],
        class_series,
    )


# ---------------------------------------------------------------------------
# JSON helper for cached redshift series
# ---------------------------------------------------------------------------
def _float_list(values) -> List[Optional[float]]:
    """Return a JSON-safe list of floats, replacing NaN/inf with None."""
    out: List[Optional[float]] = []
    for value in np.asarray(values, dtype=float):
        if np.isfinite(value):
            out.append(float(value))
        else:
            out.append(None)
    return out


def write_redshift_series_json(
    outpath: str,
    z_plot: np.ndarray,
    class_series: dict,
) -> None:
    """Save reusable per-redshift arrays for combined, major, and minor totals."""
    payload = {
        "meta": {
            "script": SCRIPT_STEM,
            "description": (
                "Direct physical row-weighted totals, averaged over Monte Carlo runs. "
                "No second multiplication by a target table is applied."
            ),
            "redshift_order": "ascending",
            "population_model_version": mps.MODEL_VERSION,
            "mass_bin_meaning": "descendant galaxy stellar mass",
            "mass_bin_edges_log10M": _float_list(EDGES),
            "M_star_TDE_Msun": M_STAR_TDE_MSUN,
            "R_star_TDE_Rsun": R_STAR_TDE_RSUN,
            "f_disk": F_DISK,
            "eta_fixed": ETA_FIXED,
            "use_mummery_mbh_dep": USE_MUMMERY_MBH_DEPENDENCE,
            "use_spin_efficiency": USE_SPIN_EFFICIENCY,
        },
        "z": _float_list(z_plot),
        "mass_bin_edges_log10M": _float_list(EDGES),
        "mass_bin_labels": list(ALL_LABELS),
        "classes": {},
    }

    for cls, data in class_series.items():
        bins = []
        for ibin, label in enumerate(ALL_LABELS):
            n_values = np.asarray(data["N_ext_by_bin"][label], dtype=float)
            e_values = np.asarray(data["E_ion_by_bin"][label], dtype=float)
            bins.append({
                "bin_label": label,
                "bin_lo_log10M": float(EDGES[ibin]),
                "bin_hi_log10M": float(EDGES[ibin + 1]),
                "N_ext_TDE_by_z": _float_list(n_values),
                "E_ion_ext_erg_by_z": _float_list(e_values),
                "N_ext_TDE_total": float(np.nan_to_num(n_values).sum()),
                "E_ion_ext_erg_total": float(np.nan_to_num(e_values).sum()),
            })
        payload["classes"][cls] = {
            "bins": bins,
            "total": {
                "N_ext_TDE_by_z": _float_list(data["overall_N_ext"]),
                "E_ion_ext_erg_by_z": _float_list(data["overall_E_ion"]),
                "N_ext_TDE_total": float(np.nan_to_num(data["overall_N_ext"]).sum()),
                "E_ion_ext_erg_total": float(np.nan_to_num(data["overall_E_ion"]).sum()),
            },
        }

    # Backward-compatible aliases for plotting helpers that expect combined data.
    payload["bins"] = payload["classes"]["combined"]["bins"]
    payload["total"] = payload["classes"]["combined"]["total"]

    with open(outpath, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
    print(f"Saved redshift-series JSON: {outpath}")


def load_redshift_series_json(path: str = SERIES_JSON_OUTPATH) -> dict:
    """Load the cached per-redshift/per-bin series JSON."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def cached_series_to_arrays(payload: dict) -> Tuple[np.ndarray, Dict[str, np.ndarray], Dict[str, np.ndarray], np.ndarray, np.ndarray]:
    """Convert a cached series-JSON payload back into NumPy arrays."""
    z_plot = np.asarray(payload["z"], dtype=float)

    series_ext_by_bin = {lbl: np.zeros_like(z_plot, dtype=float) for lbl in ALL_LABELS}
    eion_series_by_bin = {lbl: np.zeros_like(z_plot, dtype=float) for lbl in ALL_LABELS}

    for row in payload.get("bins", []):
        lbl = row.get("bin_label")
        if lbl not in series_ext_by_bin:
            continue
        series_ext_by_bin[lbl] = np.asarray(row.get("N_ext_TDE_by_z", []), dtype=float)
        eion_series_by_bin[lbl] = np.asarray(row.get("E_ion_ext_erg_by_z", []), dtype=float)

    total = payload.get("total", {})
    overall_ext = np.asarray(total.get("N_ext_TDE_by_z", np.zeros_like(z_plot)), dtype=float)
    overall_eion = np.asarray(total.get("E_ion_ext_erg_by_z", np.zeros_like(z_plot)), dtype=float)

    return z_plot, series_ext_by_bin, eion_series_by_bin, overall_ext, overall_eion


# ---------------------------------------------------------------------------
# Console summary and output table
# ---------------------------------------------------------------------------
def print_external_energy_summary(class_series: dict) -> None:
    """Print combined totals and their major/minor decomposition by bin."""
    print("\n=== External TDE and ionising-energy totals by descendant mass bin ===")
    header = (
        f"{'Bin [log10 M*]':>18}  {'N combined':>13}  {'N major':>13}  "
        f"{'N minor':>13}  {'E combined [erg]':>18}"
    )
    print(header)
    print("-" * len(header))
    for label in ALL_LABELS:
        n_comb = float(np.nan_to_num(class_series['combined']['N_ext_by_bin'][label]).sum())
        n_major = float(np.nan_to_num(class_series[mps.MERGER_MAJOR]['N_ext_by_bin'][label]).sum())
        n_minor = float(np.nan_to_num(class_series[mps.MERGER_MINOR]['N_ext_by_bin'][label]).sum())
        e_comb = float(np.nan_to_num(class_series['combined']['E_ion_by_bin'][label]).sum())
        if n_comb <= 0.0 and e_comb <= 0.0:
            continue
        print(f"{label:>18}  {n_comb:13.3e}  {n_major:13.3e}  {n_minor:13.3e}  {e_comb:18.3e}")
    print("-" * len(header))
    for cls in ('combined', mps.MERGER_MAJOR, mps.MERGER_MINOR):
        n_total = float(np.nan_to_num(class_series[cls]['overall_N_ext']).sum())
        e_total = float(np.nan_to_num(class_series[cls]['overall_E_ion']).sum())
        print(f"{cls.upper():>18}  {n_total:13.3e}  {'':13}  {'':13}  {e_total:18.3e}")


def build_external_totals_table(class_series: dict) -> pd.DataFrame:
    """Return per-bin and total combined/major/minor TDE and energy values."""
    rows = []
    for ibin, label in enumerate(ALL_LABELS):
        row = {
            "section": "external_totals",
            "bin_label": label,
            "bin_lo_log10M": float(EDGES[ibin]),
            "bin_hi_log10M": float(EDGES[ibin + 1]),
        }
        for cls, suffix in (("combined", ""), (mps.MERGER_MAJOR, "_major"),
                            (mps.MERGER_MINOR, "_minor")):
            row[f"N_ext_TDE{suffix}"] = float(
                np.nan_to_num(class_series[cls]["N_ext_by_bin"][label]).sum()
            )
            row[f"E_ion_ext_erg{suffix}"] = float(
                np.nan_to_num(class_series[cls]["E_ion_by_bin"][label]).sum()
            )
        rows.append(row)

    total = {
        "section": "external_totals",
        "bin_label": "TOTAL (by-bin sum)",
        "bin_lo_log10M": float("nan"),
        "bin_hi_log10M": float("nan"),
    }
    for cls, suffix in (("combined", ""), (mps.MERGER_MAJOR, "_major"),
                        (mps.MERGER_MINOR, "_minor")):
        total[f"N_ext_TDE{suffix}"] = float(
            np.nan_to_num(class_series[cls]["overall_N_ext"]).sum()
        )
        total[f"E_ion_ext_erg{suffix}"] = float(
            np.nan_to_num(class_series[cls]["overall_E_ion"]).sum()
        )
    rows.append(total)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Log-polynomial fit helper
# ---------------------------------------------------------------------------
def _fit_log_polynomial(
    z_centers: np.ndarray,
    y_values: np.ndarray,
    max_degree: int = 2,
    label: str = "",
) -> Tuple[Optional[np.ndarray], Optional[np.ndarray], Optional[np.ndarray]]:
    """
    Fit log10(y) as a diagnostic polynomial in z (up to max_degree), then
    return (z_line, y_line, coefficients) evaluated on a dense grid. Totals are
    computed from binned simulation outputs, not from this fit.
    Returns (None, None, None) if there are too few points to fit.
    """
    z_arr = np.asarray(z_centers, dtype=float)
    y_arr = np.asarray(y_values,  dtype=float)

    mask   = y_arr > 0.0
    n_pts  = int(mask.sum())
    if n_pts < 2:
        print(f"[fit {label}] too few positive points (N={n_pts}); skipping")
        return None, None, None

    z_fit = z_arr[mask]
    y_fit = y_arr[mask]
    degree = 1 if n_pts <= 3 else min(max_degree, n_pts - 1)

    coeffs = np.polyfit(z_fit, np.log10(y_fit), degree)
    z_line = np.linspace(z_fit.min(), z_fit.max(), 400)
    y_line = 10.0 ** np.polyval(coeffs, z_line)

    sum_data  = float(y_fit.sum())
    sum_model = float((10.0 ** np.polyval(coeffs, z_fit)).sum())
    print(f"[fit {label}] deg={degree}, N={n_pts}, sum_data={sum_data:.3e}, sum_model={sum_model:.3e}")

    return z_line, y_line, coeffs


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------
def save_figure(fig, outpath: str, dpi: int = 150) -> None:
    """Save and close a Matplotlib figure, reporting whether the file was created or overwritten."""
    existed = os.path.exists(outpath)
    fig.savefig(outpath, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    status = "Updated (overwrote)" if existed else "Saved (new)"
    print(f"{status}: {outpath}")


def plot_external_ionizing_energy(
    z_plot: np.ndarray,
    eion_series_by_bin: Dict,
    overall_eion: np.ndarray,
) -> pd.DataFrame:
    """
    Produce two figures:
      (1) Per-mass-bin grid: ionising energy vs redshift
      (2) Single-panel total across all bins

    Returns a fit-coefficient DataFrame for use in the JSON / XLSX output.
    """
    active_bins = [
        lbl for lbl in ALL_LABELS
        if np.any(np.nan_to_num(eion_series_by_bin[lbl], nan=0.0) > 0.0)
    ]
    if not active_bins:
        print("\n[plot] No external ionizing energy to plot.")
        return pd.DataFrame(columns=["section", "bin_label", "deg", "poly_form", "a0", "a1", "a2"])

    active_bins = sorted(active_bins, key=lambda lbl: EDGES[ALL_LABELS.index(lbl)])

    # Bar width derived from the redshift grid spacing
    bar_width = float(np.abs(np.median(np.diff(z_plot)))) if len(z_plot) > 1 else 0.2

    # Shared y-limits across both figures
    all_positive = []
    for lbl in active_bins:
        vals = np.nan_to_num(eion_series_by_bin[lbl], nan=0.0)
        all_positive.append(vals[vals > 0.0])
    total_vals = np.nan_to_num(overall_eion, nan=0.0)
    all_positive.append(total_vals[total_vals > 0.0])
    all_positive = [a for a in all_positive if len(a) > 0]

    if all_positive:
        concat = np.concatenate(all_positive)
        y_min = max(float(np.min(concat)) * 0.85, 1.0)
        y_max = float(np.max(concat)) * 1.15
    else:
        y_min, y_max = 1.0, 1e6

    fit_rows_print = []
    fit_rows_df    = []

    # --- Figure 1: per-bin grid ---
    n_panels = len(active_bins)
    ncols    = 2 if n_panels <= 6 else 3
    nrows    = int(np.ceil(n_panels / ncols))

    fig1, axes = plt.subplots(nrows, ncols, figsize=(7.0 * ncols, 2.6 * nrows),
                              sharex=True, sharey=True)
    axes = np.ravel(axes)

    for j, lbl in enumerate(active_bins):
        ax    = axes[j]
        y     = np.nan_to_num(eion_series_by_bin[lbl], nan=0.0)
        color = COLOR_BY_LABEL.get(lbl, "C0")

        ax.bar(z_plot, y, width=bar_width, align="center",
               color=color, edgecolor="black", linewidth=0.5, alpha=0.35, zorder=1)

        z_line, y_line, coeffs = _fit_log_polynomial(z_plot, y, max_degree=2, label=lbl)
        if z_line is not None:
            ax.plot(z_line, y_line, "-", color=color, linewidth=2.0, zorder=3)
            degree      = len(coeffs) - 1
            coeffs_asc  = coeffs[::-1]   # [a0, a1, a2, ...]
            poly_form   = ("a0 + a1 z" if degree == 1
                           else "a0 + a1 z + a2 z^2" if degree == 2
                           else f"poly deg {degree}")
            coeff_str   = ", ".join(f"a{k}={c:.3e}" for k, c in enumerate(coeffs_asc))

            i_all     = ALL_LABELS.index(lbl)
            bin_label = f"[{EDGES[i_all]:4.2f}, {EDGES[i_all+1]:4.2f}]"
            fit_rows_print.append((bin_label, degree, poly_form, coeff_str))
            fit_rows_df.append({
                "section":    "logpoly_fits",
                "bin_label":  bin_label,
                "deg":        int(degree),
                "poly_form":  poly_form,
                "a0": float(coeffs_asc[0]) if len(coeffs_asc) > 0 else float("nan"),
                "a1": float(coeffs_asc[1]) if len(coeffs_asc) > 1 else float("nan"),
                "a2": float(coeffs_asc[2]) if len(coeffs_asc) > 2 else float("nan"),
            })

        ax.set_yscale("log")
        ax.set_ylim(y_min, y_max)
        ax.set_title(f"Bin {lbl}", fontsize=15)
        ax.tick_params(axis="x", labelsize=12)
        ax.tick_params(axis="y", labelsize=13)
        ax.grid(True, linestyle=":", linewidth=0.7)

        if j % ncols == 0:
            ax.set_ylabel(r"$E_{\rm ion,ext}$ [erg]", fontsize=14)

    for k in range(n_panels, len(axes)):
        axes[k].axis("off")

    for ax in axes[:n_panels]:
        ax.set_xlim(z_plot.min() - 0.5 * bar_width, z_plot.max() + 0.5 * bar_width)

    for ax in axes[(nrows - 1) * ncols: nrows * ncols]:
        if ax.get_visible():
            ax.set_xlabel("Redshift $z$", fontsize=14)

    selection_note = " (filter: t_external_yr > 0)" if REQUIRE_EXTERNAL_WINDOW else ""

    fig1.tight_layout(rect=[0, 0.02, 1, 0.90])

    title_ax = axes[1] if n_panels > 1 else axes[0]
    title_x = 0.5 * (title_ax.get_position().x0 + title_ax.get_position().x1)

    fig1.suptitle(
        "External TDE ionizing energy vs. redshift\n"
        f"Per host stellar-mass bin{selection_note}",
        x=title_x, y=0.975, fontsize=18,
    )

    out1 = os.path.join(FIG_DIR, "external_tde_ionizing_energy_per_bin.png")
    save_figure(fig1, out1)

    # --- Figure 2: total across all bins ---
    fig2, ax = plt.subplots(figsize=(7.6, 3.2))
    ax.bar(z_plot, total_vals, width=bar_width, align="center",
           color="black", edgecolor="red", linewidth=0.8, alpha=0.4, zorder=1)

    z_line_tot, y_line_tot, coeffs_tot = _fit_log_polynomial(
        z_plot, total_vals, max_degree=2, label="TOTAL"
    )
    if z_line_tot is not None:
        ax.plot(z_line_tot, y_line_tot, "-", color="red", linewidth=2.5, zorder=3)
        degree_tot     = len(coeffs_tot) - 1
        coeffs_asc_tot = coeffs_tot[::-1]
        poly_form_tot  = ("a0 + a1 z" if degree_tot == 1
                          else "a0 + a1 z + a2 z^2" if degree_tot == 2
                          else f"poly deg {degree_tot}")
        coeff_str_tot  = ", ".join(f"a{k}={c:.3e}" for k, c in enumerate(coeffs_asc_tot))
        fit_rows_print.append(("TOTAL", degree_tot, poly_form_tot, coeff_str_tot))
        fit_rows_df.append({
            "section":   "logpoly_fits",
            "bin_label": "TOTAL",
            "deg":       int(degree_tot),
            "poly_form": poly_form_tot,
            "a0": float(coeffs_asc_tot[0]) if len(coeffs_asc_tot) > 0 else float("nan"),
            "a1": float(coeffs_asc_tot[1]) if len(coeffs_asc_tot) > 1 else float("nan"),
            "a2": float(coeffs_asc_tot[2]) if len(coeffs_asc_tot) > 2 else float("nan"),
        })

    ax.set_yscale("log")
    ax.set_ylim(y_min, y_max)
    ax.set_xlim(z_plot.min() - 0.5 * bar_width, z_plot.max() + 0.5 * bar_width)
    ax.set_title("Total external ionizing energy (all host-mass bins)", fontsize=16)
    ax.set_xlabel("Redshift $z$", fontsize=14)
    ax.set_ylabel(r"$E_{\rm ion,ext}$ [erg]", fontsize=14)
    ax.tick_params(axis="x", labelsize=12)
    ax.tick_params(axis="y", labelsize=13)
    ax.grid(True, linestyle=":", linewidth=0.7)

    fig2.tight_layout()
    out2 = os.path.join(FIG_DIR, "external_tde_ionizing_energy_total.png")
    save_figure(fig2, out2)

    # Print fit coefficient table
    if fit_rows_print:
        print("\n=== Log-polynomial fits for external ionizing-energy histories ===")
        print("log10 E_ion,ext(z) = a0 + a1*z + ... + aN*z^N\n")
        header = (
            f"{'Bin [log10 M*]':>18}  "
            f"{'deg':>3}  "
            f"{'poly form':>22}  "
            f"{'coefficients':>60}"
        )
        print(header)
        print("-" * len(header))
        for bin_label, deg, poly_form, coeff_str in fit_rows_print:
            print(f"{bin_label:>18}  {deg:3d}  {poly_form:>22}  {coeff_str:>60}")

    return pd.DataFrame(fit_rows_df)


# ---------------------------------------------------------------------------
# JSON output
# ---------------------------------------------------------------------------
def write_tables_json(
    outpath: str,
    totals_df: pd.DataFrame,
    fits_df: pd.DataFrame,
) -> None:
    """
    Write the external TDE totals and log-polynomial fit coefficients to a
    JSON file, matching the structure used by ratio_scan_vcent.py.
    """
    def _df_to_records(df: pd.DataFrame) -> list:
        """Convert a DataFrame to a list of dicts, replacing NaN with None."""
        records = []
        for row in df.to_dict(orient="records"):
            records.append({
                k: (None if (isinstance(v, float) and np.isnan(v)) else v)
                for k, v in row.items()
            })
        return records

    payload = {
        "meta": {
            "script":                  SCRIPT_STEM,
            "M_star_TDE_Msun":         M_STAR_TDE_MSUN,
            "R_star_TDE_Rsun":         R_STAR_TDE_RSUN,
            "f_disk":                  F_DISK,
            "f_ion_phase":             F_ION_PHASE,
            "eta_fixed":               ETA_FIXED,
            "use_mummery_mbh_dep":     USE_MUMMERY_MBH_DEPENDENCE,
            "use_spin_efficiency":     USE_SPIN_EFFICIENCY,
            "require_external_window": REQUIRE_EXTERNAL_WINDOW,
            "population_model_version": mps.MODEL_VERSION,
            "weighting": "direct sum(weight * quantity), then mean over runs",
            "mass_bin_meaning": "descendant galaxy stellar mass",
        },
        "external_totals": _df_to_records(totals_df) if totals_df is not None else [],
        "logpoly_fits":    _df_to_records(fits_df)   if fits_df   is not None else [],
    }

    with open(outpath, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"Saved JSON tables: {outpath}")


# ---------------------------------------------------------------------------
# XLSX output
# ---------------------------------------------------------------------------
def write_tables_xlsx(
    outpath: str,
    totals_df: pd.DataFrame,
    fits_df: pd.DataFrame,
) -> None:
    """Write external TDE totals and fit coefficients to a formatted Excel workbook."""
    if not _HAS_OPENPYXL:
        print(f"[warning] openpyxl not available; skipping XLSX output: {outpath}")
        print("          Install with: pip install openpyxl")
        return

    if (totals_df is None or totals_df.empty) and (fits_df is None or fits_df.empty):
        print(f"[warning] No tables to write to XLSX: {outpath}")
        return

    with pd.ExcelWriter(outpath, engine="openpyxl") as writer:
        if totals_df is not None and not totals_df.empty:
            totals_df.to_excel(writer, sheet_name="external_totals", index=False)
        if fits_df is not None and not fits_df.empty:
            fits_df.to_excel(writer, sheet_name="logpoly_fits", index=False)

        wb = writer.book

        def _format_sheet(ws, numfmt_by_col: dict, freeze_at: str = "A2"):
            ws.freeze_panes = freeze_at
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                f = copy(cell.font); f.bold = True; cell.font = f
                a = copy(cell.alignment); a.horizontal = "center"; cell.alignment = a

            header     = [c.value for c in ws[1]]
            col_index  = {name: j + 1 for j, name in enumerate(header) if name is not None}

            for col_name, fmt in numfmt_by_col.items():
                if col_name not in col_index:
                    continue
                cidx = col_index[col_name]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=cidx).number_format = fmt

            for j in range(1, ws.max_column + 1):
                col_letter = get_column_letter(j)
                max_len = max(
                    (len(str(ws.cell(row=r, column=j).value or ""))
                     for r in range(1, ws.max_row + 1)),
                    default=0,
                )
                ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

        if "external_totals" in wb.sheetnames:
            _format_sheet(wb["external_totals"], {
                "bin_lo_log10M": "0.00",
                "bin_hi_log10M": "0.00",
                "N_ext_TDE":             "0.00E+00",
                "E_ion_ext_erg":         "0.00E+00",
                "N_ext_TDE_major":       "0.00E+00",
                "E_ion_ext_erg_major":   "0.00E+00",
                "N_ext_TDE_minor":       "0.00E+00",
                "E_ion_ext_erg_minor":   "0.00E+00",
            })

        if "logpoly_fits" in wb.sheetnames:
            _format_sheet(wb["logpoly_fits"], {
                "a0": "0.0E+00",
                "a1": "0.0E+00",
                "a2": "0.0E+00",
            })

    print(f"Saved XLSX tables: {outpath}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    """Aggregate external TDE counts and ionising energy from simulation outputs, then write summary tables and figures."""
    print("\n=== External TDE ionizing energy budget ===\n")

    eion_ref = ionizing_energy_per_tde_erg(ETA_FIXED)
    print("Per-TDE ionizing energy parameters:")
    print(f"  M_star_TDE     = {M_STAR_TDE_MSUN:.3f} M_sun")
    print(f"  R_star_TDE     = {R_STAR_TDE_RSUN:.3f} R_sun")
    print(f"  f_disk         = {F_DISK:.3f}")
    print(f"  f_ion_phase    = {F_ION_PHASE:.3f}  (legacy fixed-DeltaM path only)")
    print(f"  eta_fixed      = {ETA_FIXED:.3f}")
    print(f"  E_ion (fixed)  = {eion_ref:.3e} erg")

    if PRINT_EQUIV_PHOTONS:
        n_equiv = eion_ref / (E_PHOTON_EFF_EV * EV_TO_ERG)
        print(f"  Rydberg-equivalent photons = E_ion / 13.6 eV = {n_equiv:.3e}")

    print("\nModel switches:")
    print(f"  USE_MUMMERY_MBH_DEPENDENCE = {USE_MUMMERY_MBH_DEPENDENCE}")
    print(f"  USE_SPIN_EFFICIENCY        = {USE_SPIN_EFFICIENCY}")
    print()

    if USE_MUMMERY_MBH_DEPENDENCE:
        print("Mummery ΔM_acc sanity check (solar-type star, M_disk = M_star_TDE * f_disk):")
        m_disk = M_STAR_TDE_MSUN * F_DISK
        for mbh_test in [1e5, 1e6, 1e7, 1e8, 2e8]:
            d_m = float(mummery_accreted_mass_msun(np.array([mbh_test]), m_disk)[0])
            e   = ETA_FIXED * (d_m * M_SUN) * C_LIGHT**2 * 1e7
            print(f"  M_BH = {mbh_test:.1e} M_sun:  DeltaM_acc = {d_m:.3f} M_sun  ->  E_ion = {e:.3e} erg")
        print()

    # Load revised simulation catalogues and aggregate direct physical weights.
    z_array, paths_by_z = discover_parquet_snapshots(PARQUET_DIR)
    (
        z_plot, series_ext_by_bin, eion_series_by_bin,
        overall_ext, overall_eion, class_series,
    ) = accumulate_external_tdes_and_energy(z_array, paths_by_z)

    # Save the reusable arrays immediately after aggregation. This JSON is the
    # lightweight cache used for later subset figures; no Parquet reads are
    # needed once this file exists.
    write_redshift_series_json(SERIES_JSON_OUTPATH, z_plot, class_series)

    print_external_energy_summary(class_series)

    total_tde  = float(np.nan_to_num(overall_ext,  nan=0.0).sum())
    total_eion = float(np.nan_to_num(overall_eion, nan=0.0).sum())

    print("\n=== Global totals (external, direct weighted mean over runs) ===")
    print(f"  Total external TDEs          = {total_tde:.3e}")
    print(f"  Total external E_ion [erg]   = {total_eion:.3e}")

    if PRINT_EQUIV_PHOTONS:
        n_equiv = total_eion / (E_PHOTON_EFF_EV * EV_TO_ERG)
        print(f"  Rydberg-equivalent photons   = {n_equiv:.3e}")

    print()

    # Build output tables and write files
    totals_df = build_external_totals_table(class_series)
    fits_df   = plot_external_ionizing_energy(z_plot, eion_series_by_bin, overall_eion)

    write_tables_json(JSON_OUTPATH, totals_df, fits_df)
    write_tables_xlsx(XLSX_OUTPATH, totals_df, fits_df)

    print(f"\n>>> {SCRIPT_STEM}.py finished successfully.")


if __name__ == "__main__":
    main()
